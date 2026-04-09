"""
output/excel_reports.py — Enterprise & Weekly Excel report generators.

Two report types:
  1. Enterprise Repository Inventory — multi-org overview with all repos + health
  2. Weekly Organisation Report — single-org deep dive with 10 detail sheets
"""

from __future__ import annotations

import logging
import math
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from config import settings
from models import AlertState, Severity

logger = logging.getLogger(__name__)


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _gen_date() -> str:
    return datetime.now(timezone.utc).strftime("%A, %d %B %Y  —  %H:%M:%S")


def _week_label() -> str:
    now = datetime.now(timezone.utc)
    return f"Week {now.isocalendar()[1]},  {now.year}"


def _age_days(dt: Optional[datetime]) -> int:
    if not dt:
        return 0
    now = datetime.now(timezone.utc)
    if dt.tzinfo is None:
        from datetime import timezone as tz
        dt = dt.replace(tzinfo=tz.utc)
    return max(0, (now - dt).days)


def _safe(val: Any, max_len: int = 200) -> str:
    s = str(val) if val else ""
    return s[:max_len] if len(s) > max_len else s


# ── Colour palette (matching main renderer) ──────────────────────────────────

_NAVY        = "1A1A2E"
_DARK_BLUE   = "16213E"
_WHITE       = "FFFFFF"
_LIGHT_GREY  = "F0F2F5"
_LIGHT_BG    = "F8FAFC"
_BORDER_GREY = "D1D5DB"
_RED         = "E63946"
_ORANGE      = "F4845F"
_AMBER       = "F59E0B"
_GREEN       = "10B981"
_BLUE        = "3B82F6"
_DARK_TEXT   = "1E293B"
_GREY_TEXT   = "64748B"
_PURPLE      = "7C3AED"

_SEVERITY_BG = {
    "critical": "FEE2E2", "high": "FFEDD5", "medium": "FEF3C7",
    "low": "D1FAE5", "warning": "FEF3C7", "note": "DBEAFE",
    "none": _LIGHT_GREY, "unknown": _LIGHT_GREY,
}
_SEVERITY_FG = {
    "critical": _RED, "high": "C2410C", "medium": "B45309",
    "low": "047857", "warning": "B45309", "note": "1D4ED8",
    "none": _GREY_TEXT, "unknown": _GREY_TEXT,
}

# Risk level thresholds
def _risk_level(critical: int, high: int, total: int) -> str:
    if critical > 10:
        return "🔴 Critical"
    if critical > 0:
        return "🔴 Critical"
    if high > 10:
        return "🟠 High"
    if high > 0:
        return "🟠 High"
    if total > 0:
        return "🟡 Medium"
    return "🟢 Clean"


def _import_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.series import DataPoint
        return openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")


# ══════════════════════════════════════════════════════════════════════════════
#  Shared Excel helpers
# ══════════════════════════════════════════════════════════════════════════════

def _styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin_border = Border(
        left=Side(style="thin", color=_BORDER_GREY), right=Side(style="thin", color=_BORDER_GREY),
        top=Side(style="thin", color=_BORDER_GREY), bottom=Side(style="thin", color=_BORDER_GREY),
    )
    hdr_fill = PatternFill("solid", fgColor=_NAVY)
    hdr_font = Font(bold=True, size=10, color=_WHITE, name="Calibri")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(size=10, color=_DARK_TEXT, name="Calibri")
    body_align = Alignment(vertical="center", wrap_text=True)
    stripe_a = PatternFill("solid", fgColor=_WHITE)
    stripe_b = PatternFill("solid", fgColor=_LIGHT_BG)
    return {
        "thin_border": thin_border, "hdr_fill": hdr_fill, "hdr_font": hdr_font,
        "hdr_align": hdr_align, "body_font": body_font, "body_align": body_align,
        "stripe_a": stripe_a, "stripe_b": stripe_b,
    }


def _apply_header_row(ws, ncols: int, row: int, st: dict):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = st["hdr_font"]
        cell.fill = st["hdr_fill"]
        cell.alignment = st["hdr_align"]
        cell.border = st["thin_border"]
    ws.row_dimensions[row].height = 30


def _apply_data_cell(cell, row_idx: int, st: dict):
    cell.font = st["body_font"]
    cell.alignment = st["body_align"]
    cell.border = st["thin_border"]
    cell.fill = st["stripe_b"] if row_idx % 2 == 0 else st["stripe_a"]


def _apply_severity(cell, severity_str: str, st: dict):
    from openpyxl.styles import Font, PatternFill, Alignment
    sev = (severity_str or "unknown").lower()
    cell.fill = PatternFill("solid", fgColor=_SEVERITY_BG.get(sev, _LIGHT_GREY))
    cell.font = Font(bold=True, size=10, color=_SEVERITY_FG.get(sev, _DARK_TEXT), name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = st["thin_border"]
    cell.value = sev.upper()


def _set_col_widths(ws, widths: dict[str, float]):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _title_block(ws, text: str, sub_text: str, ncols: int):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    end_col = get_column_letter(ncols)
    ws.merge_cells(f"A1:{end_col}1")
    c1 = ws["A1"]
    c1.value = text
    c1.font = Font(bold=True, size=16, color=_WHITE, name="Calibri")
    c1.fill = PatternFill("solid", fgColor=_NAVY)
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 44

    ws.merge_cells(f"A2:{end_col}2")
    c2 = ws["A2"]
    c2.value = sub_text
    c2.font = Font(italic=True, size=9, color=_WHITE, name="Calibri")
    c2.fill = PatternFill("solid", fgColor=_DARK_BLUE)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22


def _write_data_rows(ws, headers: list[str], rows: list[list], start_row: int,
                     st: dict, severity_col: int | None = None, auto_filter: bool = True,
                     freeze: bool = True):
    """Write header + data rows with standard styling."""
    from openpyxl.utils import get_column_letter
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    _apply_header_row(ws, len(headers), start_row, st)
    if auto_filter:
        end_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A{start_row}:{end_col}{start_row}"
    if freeze:
        ws.freeze_panes = f"A{start_row + 1}"
    for i, row_data in enumerate(rows):
        r = start_row + 1 + i
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if severity_col is not None and c == severity_col:
                _apply_severity(cell, str(val) if val else "unknown", st)
            else:
                _apply_data_cell(cell, i, st)
        ws.row_dimensions[r].height = 22


# ══════════════════════════════════════════════════════════════════════════════
#  Helper: build per-repo alert aggregation from alert lists
# ══════════════════════════════════════════════════════════════════════════════

def _aggregate_per_repo(dep_alerts, cs_alerts, ss_alerts):
    """Return dict keyed by repo full_name → {dep, cs, ss, severity counts}."""
    repo_data = defaultdict(lambda: {
        "dep_total": 0, "cs_total": 0, "ss_total": 0,
        "critical": 0, "high": 0, "medium": 0, "low": 0,
        "dep_open": 0, "cs_open": 0, "ss_open": 0,
    })

    for a in dep_alerts:
        key = a.repository.full_name
        repo_data[key]["dep_total"] += 1
        if a.state == AlertState.OPEN:
            repo_data[key]["dep_open"] += 1
            sev = a.severity.value.lower()
            if sev in repo_data[key]:
                repo_data[key][sev] += 1

    for a in cs_alerts:
        key = a.repository.full_name
        repo_data[key]["cs_total"] += 1
        if a.state == AlertState.OPEN:
            repo_data[key]["cs_open"] += 1
            sev = a.severity.value.lower()
            if sev in repo_data[key]:
                repo_data[key][sev] += 1

    for a in ss_alerts:
        key = a.repository.full_name
        repo_data[key]["ss_total"] += 1
        if a.state == AlertState.OPEN:
            repo_data[key]["ss_open"] += 1
            repo_data[key]["critical"] += 1  # secrets are always critical

    return dict(repo_data)


# ══════════════════════════════════════════════════════════════════════════════
#  REPORT TYPE 1 — Enterprise Repository Inventory
#  (multi-org, used with --list-orgs)
# ══════════════════════════════════════════════════════════════════════════════

def render_enterprise_report(
    all_org_data: list[dict[str, Any]],
    output_dir: Optional[Path] = None,
) -> Path:
    """
    Generate Enterprise Repository Inventory Excel report.

    `all_org_data` is a list of dicts, one per org:
      [{ "org": str, "repos": [raw_repo_dict, ...],
         "dependabot": [DependabotAlert, ...],
         "code_scanning": [CodeScanningAlert, ...],
         "secret_scanning": [SecretScanningAlert, ...],
         "summary": SecuritySummary }, ...]
    """
    _import_openpyxl()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint

    output_dir = output_dir or settings.OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"Enterprise_Repository_Inventory_{_timestamp()}.xlsx"

    wb = openpyxl.Workbook()
    gen = _gen_date()
    st = _styles()

    total_orgs = len(all_org_data)
    total_repos = sum(len(d.get("repos", [])) for d in all_org_data)

    # Build flat repo list with alert aggregation
    all_repo_rows = []
    health_rows = []
    pivot_rows = []

    for org_data in all_org_data:
        org_name = org_data["org"]
        repos = org_data.get("repos", [])
        per_repo = _aggregate_per_repo(
            org_data.get("dependabot", []),
            org_data.get("code_scanning", []),
            org_data.get("secret_scanning", []),
        )

        for repo in repos:
            full_name = repo.get("full_name", "")
            repo_name = repo.get("name", "")
            desc = repo.get("description") or "No description"
            language = repo.get("language") or "None"
            visibility = repo.get("visibility", "private").title()
            is_fork = "Yes" if repo.get("fork") else "No"
            size_kb = repo.get("size", 0)
            stars = repo.get("stargazers_count", 0)
            forks = repo.get("forks_count", 0)
            default_branch = repo.get("default_branch", "main")
            archived = repo.get("archived", False)
            status = "Archived" if archived else "Active"
            license_name = (repo.get("license") or {}).get("name") or "None"
            html_url = repo.get("html_url", "")
            created = (repo.get("created_at") or "")[:10]
            updated = (repo.get("updated_at") or "")[:10]
            pushed = (repo.get("pushed_at") or "")[:10]
            days_since_push = 0
            if repo.get("pushed_at"):
                try:
                    push_dt = datetime.fromisoformat(repo["pushed_at"].replace("Z", "+00:00"))
                    days_since_push = (datetime.now(timezone.utc) - push_dt).days
                except (ValueError, TypeError):
                    pass

            rd = per_repo.get(full_name, {})
            dep_alerts = rd.get("dep_open", 0)
            cs_alerts_count = rd.get("cs_open", 0)
            ss_alerts_count = rd.get("ss_open", 0)
            total_alerts = dep_alerts + cs_alerts_count + ss_alerts_count

            # All Repositories sheet row
            all_repo_rows.append([
                org_name, repo_name, full_name, _safe(desc, 100),
                org_name, language, visibility, status, is_fork,
                size_kb, stars, forks, default_branch, days_since_push,
                created, updated, pushed, license_name, html_url,
                dep_alerts, cs_alerts_count, ss_alerts_count, total_alerts,
            ])

            # Repository Health sheet row
            dep_enabled = "Enabled" if dep_alerts > 0 or repo.get("has_dependabot", True) else "Unknown"
            cs_enabled = "Enabled" if cs_alerts_count > 0 else "Unknown"
            ss_enabled = "Enabled" if ss_alerts_count > 0 else "Unknown"
            crit = rd.get("critical", 0)
            high = rd.get("high", 0)
            compliance = _calc_compliance(dep_enabled, cs_enabled, ss_enabled)
            sev_label = "Critical" if crit > 0 else ("High" if high > 0 else ("Medium" if total_alerts > 0 else "Clean"))

            health_rows.append([
                org_name, repo_name, "N/A", "N/A",
                sev_label, compliance,
                dep_enabled, dep_alerts,
                cs_enabled, cs_alerts_count,
                ss_enabled, ss_alerts_count,
                total_alerts, visibility, language, days_since_push,
                "Yes" if archived else "No", status,
            ])

            # Risk Pivot row (only repos with alerts)
            if total_alerts > 0:
                pivot_rows.append([
                    org_name, repo_name,
                    rd.get("critical", 0), rd.get("high", 0),
                    rd.get("medium", 0), rd.get("low", 0), total_alerts,
                ])

    # Sort pivot by total desc
    pivot_rows.sort(key=lambda r: r[6], reverse=True)

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 1 — Executive Summary
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = _NAVY

    ws.merge_cells("A1:D1")
    ws["A1"].value = "🏢  GitHub Enterprise - Organization Inventory"
    ws["A1"].font = Font(bold=True, size=16, color=_NAVY, name="Calibri")
    ws.row_dimensions[1].height = 36

    meta = [
        ("Report Type", "Organization & Repository Inventory"),
        ("Generated On", gen),
        ("Total Organizations", str(total_orgs)),
        ("Total Repositories", str(total_repos)),
    ]
    for i, (lbl, val) in enumerate(meta, 2):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=True, size=10, color=_GREY_TEXT, name="Calibri")
        ws.cell(row=i, column=2, value=val).font = Font(size=10, color=_DARK_TEXT, name="Calibri")
        ws.row_dimensions[i].height = 22

    # Compute totals per org
    r = 7
    ws.cell(row=r, column=1, value="📊  Organization Summary").font = Font(
        bold=True, size=13, color=_NAVY, name="Calibri")
    ws.row_dimensions[r].height = 28
    r += 1

    org_summary_headers = ["Organization", "Repositories", "Dependabot Alerts",
                           "Code Scanning Alerts", "Secret Scanning Alerts", "Total Alerts"]
    for c, h in enumerate(org_summary_headers, 1):
        ws.cell(row=r, column=c, value=h)
    _apply_header_row(ws, len(org_summary_headers), r, st)
    r += 1

    grand_dep = grand_cs = grand_ss = 0
    for org_data in all_org_data:
        s = org_data.get("summary")
        dep_o = s.open_dependabot if s else 0
        cs_o = s.open_code_scanning if s else 0
        ss_o = s.open_secret_scanning if s else 0
        total_o = dep_o + cs_o + ss_o
        grand_dep += dep_o
        grand_cs += cs_o
        grand_ss += ss_o
        row_data = [org_data["org"], len(org_data.get("repos", [])), dep_o, cs_o, ss_o, total_o]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            _apply_data_cell(cell, r, st)
            if c >= 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if c == 6 and isinstance(val, int) and val > 0:
                cell.font = Font(bold=True, size=10, color=_RED, name="Calibri")
        ws.row_dimensions[r].height = 24
        r += 1

    # Grand total row
    grand_total = grand_dep + grand_cs + grand_ss
    total_row = ["TOTAL", total_repos, grand_dep, grand_cs, grand_ss, grand_total]
    for c, val in enumerate(total_row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(bold=True, size=11, color=_WHITE, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = st["thin_border"]
    ws.row_dimensions[r].height = 28

    # Severity breakdown
    r += 2
    ws.cell(row=r, column=1, value="🔥  Severity Breakdown (All Organizations)").font = Font(
        bold=True, size=13, color=_NAVY, name="Calibri")
    ws.row_dimensions[r].height = 28
    r += 1

    all_critical = sum(d.get("summary").critical_dependabot + d.get("summary").critical_code_scanning
                       for d in all_org_data if d.get("summary"))
    all_high = sum(d.get("summary").high_dependabot + d.get("summary").high_code_scanning
                   for d in all_org_data if d.get("summary"))
    all_other = max(0, grand_total - all_critical - all_high)

    sev_headers = ["Severity", "Count", "Percentage"]
    for c, h in enumerate(sev_headers, 1):
        ws.cell(row=r, column=c, value=h)
    _apply_header_row(ws, 3, r, st)
    r += 1

    for label, count, colour in [
        ("Critical", all_critical, _RED),
        ("High", all_high, _ORANGE),
        ("Medium / Low", all_other, _AMBER),
    ]:
        pct = f"{count / grand_total * 100:.1f}%" if grand_total > 0 else "0%"
        cell_l = ws.cell(row=r, column=1, value=label)
        cell_c = ws.cell(row=r, column=2, value=count)
        cell_p = ws.cell(row=r, column=3, value=pct)
        for cell in (cell_l, cell_c, cell_p):
            _apply_data_cell(cell, r, st)
        cell_l.font = Font(bold=True, size=10, color=colour, name="Calibri")
        cell_c.font = Font(bold=True, size=10, color=colour, name="Calibri")
        cell_c.alignment = Alignment(horizontal="center", vertical="center")
        cell_p.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 24
        r += 1

    _set_col_widths(ws, {"A": 28, "B": 36, "C": 20, "D": 20})

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 2 — All Repositories
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("All Repositories")
    ws2.sheet_properties.tabColor = _BLUE
    repo_headers = [
        "Organization", "Repository Name", "Full Name", "Description", "Owner",
        "Primary Language", "Visibility", "Status", "Fork", "Size (KB)", "Stars",
        "Forks", "Default Branch", "Days Since Push", "Created", "Last Updated",
        "Last Push", "License", "URL", "Dependabot Alerts", "Code Scanning Alerts",
        "Secret Scanning Alerts", "Total Security Alerts",
    ]
    _title_block(ws2, "All Repositories", f"{total_repos} repositories across {total_orgs} organizations  |  {gen}", len(repo_headers))
    ws2.row_dimensions[3].height = 6
    _write_data_rows(ws2, repo_headers, all_repo_rows, 4, st)
    _set_col_widths(ws2, {
        "A": 22, "B": 28, "C": 36, "D": 40, "E": 22, "F": 14, "G": 12,
        "H": 10, "I": 8, "J": 10, "K": 8, "L": 8, "M": 14, "N": 14,
        "O": 12, "P": 12, "Q": 12, "R": 18, "S": 40, "T": 14, "U": 16,
        "V": 16, "W": 16,
    })

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 3 — Repository Health
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Repository Health")
    ws3.sheet_properties.tabColor = _GREEN
    health_headers = [
        "Organization", "Repository", "Admins / Owner", "Admin Email",
        "Severity", "Compliance %", "Dependabot Status", "Dependabot Alerts",
        "Code Scanning Status", "Code Scan Alerts", "Secret Scanning Status",
        "Secret Alerts", "Total Alerts", "Visibility", "Language",
        "Days Since Push", "Archived", "Status",
    ]
    # Sort health rows by total alerts descending
    health_rows.sort(key=lambda r: r[12], reverse=True)
    _title_block(ws3, "Repository Health", f"{total_repos} repositories  |  {gen}", len(health_headers))
    ws3.row_dimensions[3].height = 6
    _write_data_rows(ws3, health_headers, health_rows, 4, st, severity_col=5)
    _set_col_widths(ws3, {
        "A": 22, "B": 28, "C": 14, "D": 14, "E": 12, "F": 14, "G": 16,
        "H": 14, "I": 16, "J": 14, "K": 16, "L": 12, "M": 12, "N": 12,
        "O": 14, "P": 14, "Q": 10, "R": 10,
    })

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 4 — Organization Risk Pivot
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Organization Risk Pivot")
    ws4.sheet_properties.tabColor = _RED
    pivot_headers = ["Organization", "Repository", "Critical", "High", "Medium", "Low", "Total"]
    _title_block(ws4, "Organization Risk Pivot — Open Vulnerabilities by Severity",
                 f"{len(pivot_rows)} repositories with alerts  |  {gen}", len(pivot_headers))
    ws4.row_dimensions[3].height = 6
    _write_data_rows(ws4, pivot_headers, pivot_rows, 4, st)

    # Apply conditional formatting to severity count columns
    for i, row_data in enumerate(pivot_rows):
        r = 5 + i
        crit_val = row_data[2]
        high_val = row_data[3]
        if crit_val > 0:
            cell = ws4.cell(row=r, column=3)
            cell.font = Font(bold=True, size=10, color=_RED, name="Calibri")
        if high_val > 0:
            cell = ws4.cell(row=r, column=4)
            cell.font = Font(bold=True, size=10, color="C2410C", name="Calibri")

    _set_col_widths(ws4, {"A": 22, "B": 30, "C": 12, "D": 10, "E": 12, "F": 10, "G": 10})

    # ── Save ─────────────────────────────────────────────────────────────
    wb.active = 0
    wb.save(str(path))
    logger.info("Enterprise report written: %s", path)
    return path


def _calc_compliance(dep_status: str, cs_status: str, ss_status: str) -> int:
    """Simple compliance % based on scanning feature enablement."""
    score = 0
    if dep_status == "Enabled":
        score += 40
    if cs_status == "Enabled":
        score += 40
    if ss_status == "Enabled":
        score += 20
    return score


# ══════════════════════════════════════════════════════════════════════════════
#  REPORT TYPE 2 — Weekly Organisation Report
#  (single-org deep dive with 10 sheets)
# ══════════════════════════════════════════════════════════════════════════════

def render_weekly_report(
    context: dict[str, Any],
    repos: list[dict],
    output_dir: Optional[Path] = None,
) -> Path:
    """
    Generate Weekly Organisation Security Report in Excel.

    `context` is the standard alert context dict with keys:
      org, dependabot, code_scanning, secret_scanning, summary
    `repos` is the raw repo list from GitHubClient.list_repos()
    """
    _import_openpyxl()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint

    output_dir = output_dir or settings.OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    org_name = context.get("org", "unknown")
    path = output_dir / f"{org_name}_Weekly_Report_{_week_label().replace(' ', '').replace(',', '')}_{_timestamp()}.xlsx"

    wb = openpyxl.Workbook()
    gen = _gen_date()
    week = _week_label()
    stl = _styles()
    summary = context.get("summary")
    dep_alerts = context.get("dependabot", [])
    cs_alerts = context.get("code_scanning", [])
    ss_alerts = context.get("secret_scanning", [])
    per_repo = _aggregate_per_repo(dep_alerts, cs_alerts, ss_alerts)

    total_open = (summary.open_dependabot + summary.open_code_scanning + summary.open_secret_scanning) if summary else 0
    total_crit = (summary.critical_dependabot + summary.critical_code_scanning) if summary else 0
    total_high = (summary.high_dependabot + summary.high_code_scanning) if summary else 0

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 1 — Executive Summary
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = _NAVY

    ws.merge_cells("A1:D1")
    ws["A1"].value = "🛡  GitHub Security Reporter"
    ws["A1"].font = Font(bold=True, size=16, color=_NAVY, name="Calibri")
    ws.row_dimensions[1].height = 36

    meta_rows = [
        ("Report Type", "Weekly Security Report"),
        ("Generated On", gen),
        ("Report Period", week),
    ]
    for i, (lbl, val) in enumerate(meta_rows, 2):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=True, size=10, color=_GREY_TEXT, name="Calibri")
        ws.cell(row=i, column=2, value=val).font = Font(size=10, color=_DARK_TEXT, name="Calibri")
        ws.row_dimensions[i].height = 22

    r = 6
    ws.cell(row=r, column=1, value="🏢  Organization Details").font = Font(
        bold=True, size=13, color=_NAVY, name="Calibri")
    ws.row_dimensions[r].height = 28
    r += 1

    org_meta = [
        ("Organization", org_name),
        ("Total Repositories", str(len(repos))),
        ("Repos with Alerts", str(len(per_repo))),
        ("Push Protection Bypasses", str(summary.push_protection_bypassed if summary else 0)),
    ]
    for lbl, val in org_meta:
        ws.cell(row=r, column=1, value=lbl).font = Font(bold=True, size=10, color=_GREY_TEXT, name="Calibri")
        ws.cell(row=r, column=2, value=val).font = Font(size=10, color=_DARK_TEXT, name="Calibri")
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="📊  Security KPIs").font = Font(
        bold=True, size=13, color=_NAVY, name="Calibri")
    ws.row_dimensions[r].height = 28
    r += 1

    kpi_headers = ["Metric", "Count"]
    for c, h in enumerate(kpi_headers, 1):
        ws.cell(row=r, column=c, value=h)
    _apply_header_row(ws, 2, r, stl)
    r += 1

    kpis = [
        ("Total Open Alerts", total_open, _RED if total_open > 0 else _GREEN),
        ("Critical Severity", total_crit, _RED if total_crit > 0 else _GREEN),
        ("High Severity", total_high, _ORANGE if total_high > 0 else _GREEN),
        ("Dependabot Open", summary.open_dependabot if summary else 0, _DARK_TEXT),
        ("Code Scanning Open", summary.open_code_scanning if summary else 0, _DARK_TEXT),
        ("Secret Scanning Open", summary.open_secret_scanning if summary else 0, _DARK_TEXT),
    ]
    for label, count, colour in kpis:
        cell_l = ws.cell(row=r, column=1, value=label)
        cell_c = ws.cell(row=r, column=2, value=count)
        _apply_data_cell(cell_l, r, stl)
        _apply_data_cell(cell_c, r, stl)
        cell_c.font = Font(bold=True, size=11, color=colour, name="Calibri")
        cell_c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 24
        r += 1

    _set_col_widths(ws, {"A": 28, "B": 36, "C": 20, "D": 20})

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 2 — Analysis & Progress
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Analysis & Progress")
    ws2.sheet_properties.tabColor = _BLUE
    _title_block(ws2, "📈  Trend Analysis", f"{org_name}  |  {week}", 7)
    ws2.cell(row=4, column=1,
             value="Trend data not available — requires historical snapshots from previous runs.").font = Font(
        italic=True, size=10, color=_GREY_TEXT, name="Calibri")
    ws2.row_dimensions[4].height = 24

    # Category breakdown mini-table
    r = 6
    ws2.cell(row=r, column=1, value="📊  Current Week Snapshot").font = Font(
        bold=True, size=13, color=_NAVY, name="Calibri")
    ws2.row_dimensions[r].height = 28
    r += 1
    snap_headers = ["Category", "Total", "Open", "Critical", "High", "Medium + Low"]
    snap_data = []
    if summary:
        dep_other = max(0, summary.open_dependabot - summary.critical_dependabot - summary.high_dependabot)
        cs_other = max(0, summary.open_code_scanning - summary.critical_code_scanning - summary.high_code_scanning)
        snap_data = [
            ["Dependabot", summary.total_dependabot, summary.open_dependabot,
             summary.critical_dependabot, summary.high_dependabot, dep_other],
            ["Code Scanning", summary.total_code_scanning, summary.open_code_scanning,
             summary.critical_code_scanning, summary.high_code_scanning, cs_other],
            ["Secret Scanning", summary.total_secret_scanning, summary.open_secret_scanning,
             "—", "—", "—"],
        ]
    _write_data_rows(ws2, snap_headers, snap_data, r, stl, auto_filter=False, freeze=False)
    _set_col_widths(ws2, {"A": 20, "B": 12, "C": 12, "D": 12, "E": 12, "F": 14, "G": 14})

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 3 — Top Risks
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Top Risks")
    ws3.sheet_properties.tabColor = _RED
    _title_block(ws3, "🔥  Top Security Risks", f"Top 50 critical alerts  |  {org_name}  |  {gen}", 8)
    ws3.row_dimensions[3].height = 6

    top_risk_headers = ["Type", "Repository", "Severity", "Package", "CVE", "Summary", "Age (days)", "URL"]
    top_risk_rows = []

    # Collect critical + high alerts across all types
    for a in dep_alerts:
        if a.state != AlertState.OPEN:
            continue
        sev = a.severity.value.lower()
        if sev in ("critical", "high"):
            top_risk_rows.append([
                "Dependency", a.repository.name, sev,
                a.package.name, a.advisory.cve_id or a.advisory.ghsa_id or "",
                _safe(a.advisory.summary, 80), _age_days(a.created_at),
                a.html_url or "",
            ])
    for a in cs_alerts:
        if a.state != AlertState.OPEN:
            continue
        sev = a.severity.value.lower()
        if sev in ("critical", "high"):
            top_risk_rows.append([
                "Code Scanning", a.repository.name, sev,
                a.rule.name or a.rule.id, "", _safe(a.message, 80),
                _age_days(a.created_at), a.html_url or "",
            ])
    for a in ss_alerts:
        if a.state != AlertState.OPEN:
            continue
        top_risk_rows.append([
            "Secret", a.repository.name, "critical",
            a.secret_type_display_name or a.secret_type, "",
            "Exposed secret — rotate immediately",
            _age_days(a.created_at), a.html_url or "",
        ])

    # Sort by: severity (critical first), then age (oldest first)
    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    top_risk_rows.sort(key=lambda r: (sev_order.get(r[2], 9), -r[6]))
    top_risk_rows = top_risk_rows[:50]

    _write_data_rows(ws3, top_risk_headers, top_risk_rows, 4, stl, severity_col=3)
    _set_col_widths(ws3, {"A": 14, "B": 28, "C": 12, "D": 22, "E": 18, "F": 50, "G": 10, "H": 46})

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 4 — Repository Health
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Repository Health")
    ws4.sheet_properties.tabColor = _GREEN
    rh_headers = [
        "Repository", "Owner", "Admins", "Admin Emails", "Compliance %",
        "Dependabot Status", "Dependabot Alerts", "Code Scanning Status",
        "Code Scan Alerts", "Secret Scanning Status", "Secret Alerts",
        "Branch Protection", "Security Policy", "Total Alerts",
        "Visibility", "Language", "Days Since Push", "Archived",
    ]
    rh_rows = []
    for repo in repos:
        full_name = repo.get("full_name", "")
        repo_name = repo.get("name", "")
        rd = per_repo.get(full_name, {})
        dep_c = rd.get("dep_open", 0)
        cs_c = rd.get("cs_open", 0)
        ss_c = rd.get("ss_open", 0)
        total_a = dep_c + cs_c + ss_c
        dep_st = "Enabled" if dep_c > 0 or True else "Unknown"
        cs_st = "Enabled" if cs_c > 0 else "Unknown"
        ss_st = "Enabled" if ss_c > 0 else "Unknown"
        compliance = _calc_compliance(dep_st, cs_st, ss_st)
        days_push = 0
        if repo.get("pushed_at"):
            try:
                push_dt = datetime.fromisoformat(repo["pushed_at"].replace("Z", "+00:00"))
                days_push = (datetime.now(timezone.utc) - push_dt).days
            except (ValueError, TypeError):
                pass
        has_bp = str(repo.get("has_branch_protection", False))
        has_sp = str(bool(repo.get("security_and_analysis")))

        rh_rows.append([
            repo_name, org_name, "N/A", "N/A", compliance,
            dep_st, dep_c, cs_st, cs_c, ss_st, ss_c,
            has_bp, has_sp, total_a,
            repo.get("visibility", "private"), repo.get("language") or "None",
            days_push, str(repo.get("archived", False)),
        ])

    rh_rows.sort(key=lambda r: r[13], reverse=True)

    _title_block(ws4, "Repository Health", f"{len(repos)} repositories  |  {org_name}  |  {gen}", len(rh_headers))
    ws4.row_dimensions[3].height = 6
    _write_data_rows(ws4, rh_headers, rh_rows, 4, stl)
    _set_col_widths(ws4, {
        "A": 30, "B": 20, "C": 10, "D": 14, "E": 14, "F": 16, "G": 14,
        "H": 16, "I": 14, "J": 16, "K": 12, "L": 16, "M": 16, "N": 12,
        "O": 12, "P": 14, "Q": 14, "R": 10,
    })

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 5 — Recommendations
    # ══════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Recommendations")
    ws5.sheet_properties.tabColor = _AMBER
    _title_block(ws5, "📋  Recommendations", f"{org_name}  |  {gen}", 5)
    ws5.row_dimensions[3].height = 6

    rec_headers = ["Priority", "Area", "Recommendation", "Owner", "Due Date"]
    rec_rows = _generate_recommendations(summary, repos, per_repo)
    _write_data_rows(ws5, rec_headers, rec_rows, 4, stl, auto_filter=False, freeze=False)

    # Colour-code priority column
    for i, row_data in enumerate(rec_rows):
        r = 5 + i
        cell = ws5.cell(row=r, column=1)
        pri = str(row_data[0]).lower()
        if "urgent" in pri:
            cell.font = Font(bold=True, size=10, color=_RED, name="Calibri")
        elif "high" in pri:
            cell.font = Font(bold=True, size=10, color=_ORANGE, name="Calibri")
        elif "medium" in pri:
            cell.font = Font(bold=True, size=10, color=_AMBER, name="Calibri")

    _set_col_widths(ws5, {"A": 16, "B": 24, "C": 60, "D": 24, "E": 18})

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 6 — Repository Risk Pivot
    # ══════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Repository Risk Pivot")
    ws6.sheet_properties.tabColor = _PURPLE
    _title_block(ws6, "📊  Repository Risk Pivot — Open Vulnerabilities by Severity",
                 f"{org_name}  |  {gen}", 7)
    ws6.row_dimensions[3].height = 6

    pivot_headers = ["Repository", "Critical", "High", "Medium", "Low", "Total", "Risk Level"]
    pivot_data = []
    for full_name, rd in per_repo.items():
        repo_short = full_name.split("/")[-1] if "/" in full_name else full_name
        crit = rd.get("critical", 0)
        high = rd.get("high", 0)
        med = rd.get("medium", 0)
        low = rd.get("low", 0)
        total = crit + high + med + low
        if total > 0:
            pivot_data.append([repo_short, crit, high, med, low, total, _risk_level(crit, high, total)])
    pivot_data.sort(key=lambda r: r[5], reverse=True)

    _write_data_rows(ws6, pivot_headers, pivot_data, 4, stl)
    for i, row_data in enumerate(pivot_data):
        r = 5 + i
        if row_data[1] > 0:  # critical
            ws6.cell(row=r, column=2).font = Font(bold=True, size=10, color=_RED, name="Calibri")
        if row_data[2] > 0:  # high
            ws6.cell(row=r, column=3).font = Font(bold=True, size=10, color="C2410C", name="Calibri")

    _set_col_widths(ws6, {"A": 34, "B": 12, "C": 10, "D": 12, "E": 10, "F": 10, "G": 16})

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 7 — Dependabot Details
    # ══════════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Dependabot Details")
    ws7.sheet_properties.tabColor = _RED
    _title_block(ws7, "Dependabot Vulnerability Alerts",
                 f"{len(dep_alerts)} alerts  |  {org_name}  |  {gen}", 9)
    ws7.row_dimensions[3].height = 6

    dep_headers = ["repository", "state", "severity", "package_name",
                   "package_ecosystem", "cve_id", "summary", "age_days", "url"]
    dep_rows = []
    for a in dep_alerts:
        dep_rows.append([
            a.repository.name, a.state.value, a.severity.value,
            a.package.name, a.package.ecosystem,
            a.advisory.cve_id or a.advisory.ghsa_id or "",
            _safe(a.advisory.summary, 100), _age_days(a.created_at),
            a.html_url or "",
        ])
    dep_rows.sort(key=lambda r: ({"critical": 0, "high": 1, "medium": 2, "low": 3}.get(r[2], 9), -r[7]))

    _write_data_rows(ws7, dep_headers, dep_rows, 4, stl, severity_col=3)
    _set_col_widths(ws7, {
        "A": 30, "B": 10, "C": 12, "D": 22, "E": 14, "F": 18, "G": 50, "H": 10, "I": 46,
    })

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 8 — Code Scanning Details
    # ══════════════════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("Code Scanning Details")
    ws8.sheet_properties.tabColor = _ORANGE
    _title_block(ws8, "Code Scanning Alerts (SAST)",
                 f"{len(cs_alerts)} alerts  |  {org_name}  |  {gen}", 8)
    ws8.row_dimensions[3].height = 6

    cs_headers = ["repository", "state", "security_severity_level",
                  "rule_description", "tool_name", "file_path", "age_days", "url"]
    cs_rows = []
    for a in cs_alerts:
        loc_path = a.location.path if a.location else ""
        cs_rows.append([
            a.repository.name, a.state.value, a.severity.value,
            a.rule.name or a.rule.id, a.tool_name or "",
            loc_path, _age_days(a.created_at), a.html_url or "",
        ])
    cs_rows.sort(key=lambda r: ({"critical": 0, "high": 1, "medium": 2, "low": 3}.get(r[2], 9), -r[6]))

    _write_data_rows(ws8, cs_headers, cs_rows, 4, stl, severity_col=3)
    _set_col_widths(ws8, {
        "A": 30, "B": 10, "C": 14, "D": 36, "E": 14, "F": 40, "G": 10, "H": 46,
    })

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 9 — Secret Scanning Details
    # ══════════════════════════════════════════════════════════════════════
    ws9 = wb.create_sheet("Secret Scanning Details")
    ws9.sheet_properties.tabColor = _PURPLE
    _title_block(ws9, "Secret Scanning Alerts",
                 f"{len(ss_alerts)} alerts  |  {org_name}  |  {gen}", 7)
    ws9.row_dimensions[3].height = 6

    ss_headers = ["repository", "state", "secret_type", "resolution",
                  "push_protection_bypassed", "age_days", "url"]
    ss_rows = []
    for a in ss_alerts:
        ss_rows.append([
            a.repository.name, a.state.value,
            a.secret_type_display_name or a.secret_type or "",
            a.resolved_reason or "", str(a.push_protection_bypassed),
            _age_days(a.created_at), a.html_url or "",
        ])

    _write_data_rows(ws9, ss_headers, ss_rows, 4, stl)

    # Highlight push-protection bypassed rows
    for i, row_data in enumerate(ss_rows):
        r = 5 + i
        if row_data[4] == "True":
            cell = ws9.cell(row=r, column=5)
            cell.font = Font(bold=True, size=10, color=_RED, name="Calibri")
            cell.fill = PatternFill("solid", fgColor="FEE2E2")

    _set_col_widths(ws9, {"A": 30, "B": 10, "C": 30, "D": 16, "E": 22, "F": 10, "G": 46})

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 10 — Supply Chain
    # ══════════════════════════════════════════════════════════════════════
    ws10 = wb.create_sheet("Supply Chain")
    ws10.sheet_properties.tabColor = _GREEN
    _title_block(ws10, "Supply Chain Overview",
                 f"{len(repos)} repositories  |  {org_name}  |  {gen}", 5)
    ws10.row_dimensions[3].height = 6

    sc_headers = ["Repository", "Dependency Review", "Has Dependency Files",
                  "Dependency Graph", "Total Dependencies"]
    sc_rows = []
    for repo in repos:
        repo_name = repo.get("name", "")
        # We don't have dependency graph API data, use available metadata
        has_dep_files = "Yes" if per_repo.get(repo.get("full_name", ""), {}).get("dep_total", 0) > 0 else "No"
        sc_rows.append([
            repo_name, "No", has_dep_files, "Yes", 0,
        ])
    sc_rows.sort(key=lambda r: r[0])

    _write_data_rows(ws10, sc_headers, sc_rows, 4, stl, auto_filter=False, freeze=True)
    _set_col_widths(ws10, {"A": 34, "B": 18, "C": 20, "D": 18, "E": 18})

    # ── Save ─────────────────────────────────────────────────────────────
    wb.active = 0
    wb.save(str(path))
    logger.info("Weekly report written: %s", path)
    return path


def _generate_recommendations(summary, repos: list[dict], per_repo: dict) -> list[list]:
    """Generate actionable recommendations based on current security data."""
    recs = []

    if summary:
        crit = summary.critical_dependabot + summary.critical_code_scanning
        if crit > 0:
            recs.append([
                "🔴 URGENT", "Critical Vulnerabilities",
                f"Address {crit} critical vulnerabilities immediately.",
                "Security Team + Dev Teams", "Within 48 hours",
            ])

        if summary.open_secret_scanning > 0:
            recs.append([
                "🔴 URGENT", "Secret Exposure",
                f"Revoke and rotate {summary.open_secret_scanning} exposed secrets.",
                "Security Team", "Within 24 hours",
            ])

        high = summary.high_dependabot + summary.high_code_scanning
        if high > 0:
            recs.append([
                "🟠 HIGH", "High Severity Vulnerabilities",
                f"Create remediation plan for {high} high-severity issues.",
                "Dev Team Leads", "Within 7 days",
            ])

        if summary.push_protection_bypassed > 0:
            recs.append([
                "🔴 URGENT", "Push Protection Bypasses",
                f"Investigate {summary.push_protection_bypassed} push protection bypass(es). "
                "Review bypass justifications and revoke secrets if needed.",
                "Security Team", "Within 24 hours",
            ])

    # Check scanning enablement gaps
    total_repos = len(repos)
    if total_repos > 0:
        repos_with_cs = sum(1 for fn, rd in per_repo.items() if rd.get("cs_total", 0) > 0)
        repos_with_ss = sum(1 for fn, rd in per_repo.items() if rd.get("ss_total", 0) > 0)
        cs_gap = total_repos - repos_with_cs
        ss_gap = total_repos - repos_with_ss

        if cs_gap > total_repos * 0.2:
            recs.append([
                "🟡 MEDIUM", "Code Scanning Coverage",
                f"Enable Code Scanning on {cs_gap} repos to reach 80%+ coverage.",
                "Platform Team", "Within 14 days",
            ])
        if ss_gap > total_repos * 0.2:
            recs.append([
                "🟡 MEDIUM", "Secret Scanning Coverage",
                f"Enable Secret Scanning on {ss_gap} repos to reach 80%+ coverage.",
                "Platform Team", "Within 14 days",
            ])

    if not recs:
        recs.append([
            "🟢 INFO", "Security Posture",
            "No critical findings. Continue monitoring and maintaining current security practices.",
            "Security Team", "Ongoing",
        ])

    return recs
