"""
output/renderer.py — Multi-format output renderer.
Supports: text (terminal), PDF (reportlab), HTML, Excel (openpyxl).
"""

from __future__ import annotations

import html as _html
import logging
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Literal, Optional

from config import settings

logger = logging.getLogger(__name__)

OutputFormat = Literal["text", "pdf", "html", "excel"]
ReportTemplate = Literal["executive", "technical", "compliance"]

# ── Template definitions ─────────────────────────────────────────────────────
TEMPLATES: dict[str, dict] = {
    "executive": {
        "label": "Executive Summary",
        "description": "Clean, high-level report for leadership. KPIs and AI analysis — no raw alert tables.",
        "show_alert_tables": False,
        "show_summary_table": True,
        "show_top_repos": True,
        "excel_alert_sheets": False,
    },
    "technical": {
        "label": "Technical Detail",
        "description": "Full detail with all alert tables, code locations, severity breakdowns.",
        "show_alert_tables": True,
        "show_summary_table": True,
        "show_top_repos": True,
        "excel_alert_sheets": True,
    },
    "compliance": {
        "label": "Compliance / Audit",
        "description": "Formal report with classification banner, numbered sections, and sign-off block.",
        "show_alert_tables": True,
        "show_summary_table": True,
        "show_top_repos": True,
        "excel_alert_sheets": True,
    },
}


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _gen_date() -> str:
    return datetime.now(timezone.utc).strftime("%B %d, %Y at %H:%M UTC")


def _safe_html(text: str) -> str:
    return _html.escape(text, quote=True)


def _sanitise_llm_html(text: str) -> str:
    """Strip dangerous HTML tags from LLM output before rendering."""
    text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<(iframe|object|embed|form|input|link|meta|style)[^>]*/?>', '', text, flags=re.IGNORECASE)
    text = re.sub(r'</(iframe|object|embed|form|input|link|meta|style)>', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\bon\w+\s*=\s*["\'][^"\']*["\']', '', text, flags=re.IGNORECASE)
    return text


def _content_to_html(content: str) -> str:
    """Convert LLM plain-text output into styled HTML with section awareness."""
    content = _sanitise_llm_html(content)
    lines = content.split("\n")
    html_parts: list[str] = []
    in_list = False

    for line in lines:
        stripped = line.strip()
        safe = _safe_html(stripped)

        if not stripped:
            if in_list:
                html_parts.append("</ul>")
                in_list = False
            html_parts.append("<br>")
            continue

        # Markdown headings
        if stripped.startswith("### "):
            if in_list:
                html_parts.append("</ul>"); in_list = False
            html_parts.append(f'<h4 class="section-h4">{_safe_html(stripped[4:])}</h4>')
        elif stripped.startswith("## "):
            if in_list:
                html_parts.append("</ul>"); in_list = False
            html_parts.append(f'<h3 class="section-h3">{_safe_html(stripped[3:])}</h3>')
        elif stripped.startswith("# "):
            if in_list:
                html_parts.append("</ul>"); in_list = False
            html_parts.append(f'<h2 class="section-h2">{_safe_html(stripped[2:])}</h2>')
        # ALL-CAPS headings or separator lines
        elif re.match(r"^[=\-]{3,}$", stripped):
            if in_list:
                html_parts.append("</ul>"); in_list = False
            html_parts.append('<hr class="section-rule">')
        elif stripped.isupper() and len(stripped) > 5 and not stripped.startswith("-"):
            if in_list:
                html_parts.append("</ul>"); in_list = False
            html_parts.append(f'<h3 class="section-h3">{safe}</h3>')
        # Numbered items
        elif re.match(r"^\d+[\.\)]\s", stripped):
            if in_list:
                html_parts.append("</ul>"); in_list = False
            item_text = re.sub(r"^\d+[\.\)]\s*", "", safe)
            # Bold the first few words if followed by colon or dash
            item_text = re.sub(r"^(\*\*.*?\*\*)", lambda m: f"<strong>{_safe_html(m.group(1)[2:-2])}</strong>", item_text)
            item_text = re.sub(r"^([^:—\-]+[:—\-])", lambda m: f"<strong>{m.group(1)}</strong>", item_text, count=1)
            html_parts.append(f'<div class="numbered-item">{safe}</div>')
        # Bullet items
        elif stripped.startswith("- ") or stripped.startswith("* "):
            if not in_list:
                html_parts.append('<ul class="bullet-list">'); in_list = True
            item_text = _safe_html(stripped[2:])
            item_text = re.sub(r"^([^:—\-]+[:—\-])", lambda m: f"<strong>{m.group(1)}</strong>", item_text, count=1)
            html_parts.append(f"<li>{item_text}</li>")
        # Bold markers
        elif "**" in stripped:
            text = safe
            text = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", text)
            html_parts.append(f"<p>{text}</p>")
        else:
            html_parts.append(f"<p>{safe}</p>")

    if in_list:
        html_parts.append("</ul>")

    return "\n    ".join(html_parts)


def _compute_severity_breakdown(context: dict[str, Any]) -> dict[str, int]:
    """Return severity counts from raw alert objects."""
    counts: Counter = Counter()
    from models import AlertState
    for a in context.get("dependabot", []):
        if a.state == AlertState.OPEN:
            counts[a.severity.value] += 1
    for a in context.get("code_scanning", []):
        if a.state == AlertState.OPEN:
            counts[a.severity.value] += 1
    for a in context.get("secret_scanning", []):
        if a.state == AlertState.OPEN:
            counts["critical"] += 1  # secrets are always critical
    return dict(counts)


def _compliance_signoff_html(gen_date: str) -> str:
    """Return the HTML sign-off block for compliance template."""
    return f"""<div class="signoff">
      <h3>Report Sign-Off</h3>
      <div class="signoff-grid">
        <div class="label">Prepared by:</div><div class="value">AI Git Guard — Automated Security Analysis</div>
        <div class="label">Date:</div><div class="value">{_safe_html(gen_date)}</div>
        <div class="label">Classification:</div><div class="value" style="color:var(--red);font-weight:700">CONFIDENTIAL — FOR INTERNAL USE ONLY</div>
        <div class="label">Reviewed by:</div><div class="value">&nbsp;</div>
        <div class="label">Approved by:</div><div class="value">&nbsp;</div>
      </div>
    </div>"""


def render(
    content: str,
    title: str = "AI Git Guard Report",
    fmt: Optional[OutputFormat] = None,
    output_dir: Optional[Path] = None,
    context: Optional[dict[str, Any]] = None,
    template: ReportTemplate = "technical",
) -> Path | str:
    """
    Render `content` in the requested format and template style.
    Returns the file path (for file formats) or the formatted string (for 'text').
    `context` is the full alert context dict (used by Excel/HTML for structured data).
    `template` selects the visual template: executive | technical | compliance.
    """
    fmt = fmt or settings.DEFAULT_OUTPUT_FORMAT  # type: ignore[assignment]
    output_dir = output_dir or settings.OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    tcfg = TEMPLATES.get(template, TEMPLATES["technical"])

    filename = f"ghas_report_{_timestamp()}"

    # Derive a meaningful filename slug from the title
    _slug = re.sub(r"AI Git Guard[^a-zA-Z]*", "", title).strip()
    _slug = re.sub(r"[^a-zA-Z0-9]+", "_", _slug).strip("_").lower()
    if _slug:
        filename = f"{_slug}_{_timestamp()}"

    if fmt == "text":
        return _render_text(content, title)
    elif fmt == "pdf":
        return _render_pdf(content, title, output_dir / f"{filename}.pdf", context, template, tcfg)
    elif fmt == "html":
        return _render_html(content, title, output_dir / f"{filename}.html", context, template, tcfg)
    elif fmt == "excel":
        return _render_excel(content, title, output_dir / f"{filename}.xlsx", context, template, tcfg)
    else:
        raise ValueError(f"Unknown output format: '{fmt}'. Choose: text | pdf | html | excel")


# ── Text ─────────────────────────────────────────────────────────────────────

def _render_text(content: str, title: str) -> str:
    border = "=" * 70
    header = f"\n{border}\n  {title}\n  Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n{border}\n"
    return header + content + f"\n{border}\n"


# ══════════════════════════════════════════════════════════════════════════════
#  PDF
# ══════════════════════════════════════════════════════════════════════════════

def _render_pdf(content: str, title: str, path: Path, context: Optional[dict[str, Any]] = None,
                template: ReportTemplate = "technical", tcfg: dict | None = None) -> Path:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.lib.colors import HexColor, white, black
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table as RLTable,
            TableStyle, HRFlowable, KeepTogether,
        )
    except ImportError:
        raise RuntimeError("reportlab is not installed. Run: pip install reportlab")

    # Colours
    NAVY      = HexColor("#1a1a2e")
    DARK_BLUE = HexColor("#16213e")
    RED       = HexColor("#e63946")
    ORANGE    = HexColor("#f4845f")
    AMBER     = HexColor("#f59e0b")
    GREEN     = HexColor("#10b981")
    BLUE      = HexColor("#3b82f6")
    GREY      = HexColor("#64748b")
    LIGHT_BG  = HexColor("#f8fafc")
    BORDER    = HexColor("#e2e8f0")

    page_w, page_h = A4
    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()

    # Custom styles
    s_title = ParagraphStyle("RptTitle", parent=styles["Heading1"],
        fontSize=14, textColor=white, spaceAfter=2, fontName="Helvetica-Bold",
        alignment=TA_CENTER)
    s_subtitle = ParagraphStyle("RptSub", parent=styles["Normal"],
        fontSize=8, textColor=HexColor("#94a3b8"), spaceAfter=0, fontName="Helvetica",
        alignment=TA_CENTER)
    s_h2 = ParagraphStyle("RptH2", parent=styles["Heading2"],
        fontSize=13, textColor=NAVY, spaceBefore=16, spaceAfter=8,
        borderPadding=(0, 0, 4, 0), fontName="Helvetica-Bold")
    s_h3 = ParagraphStyle("RptH3", parent=styles["Heading3"],
        fontSize=11, textColor=DARK_BLUE, spaceBefore=12, spaceAfter=6, fontName="Helvetica-Bold")
    s_body = ParagraphStyle("RptBody", parent=styles["Normal"],
        fontSize=9.5, leading=14, textColor=black, spaceAfter=4, fontName="Helvetica")
    s_bullet = ParagraphStyle("RptBullet", parent=s_body,
        leftIndent=16, bulletIndent=6, spaceBefore=2, spaceAfter=2)
    s_num = ParagraphStyle("RptNum", parent=s_body,
        leftIndent=16, spaceBefore=3, spaceAfter=3)
    s_kpi_val = ParagraphStyle("KpiVal", fontSize=20, leading=24, alignment=TA_CENTER,
        textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0, spaceBefore=0)
    s_kpi_lbl = ParagraphStyle("KpiLbl", fontSize=7, leading=9, alignment=TA_CENTER,
        textColor=GREY, fontName="Helvetica", spaceBefore=2, spaceAfter=0)
    s_footer = ParagraphStyle("Footer", parent=styles["Normal"],
        fontSize=7, textColor=GREY, alignment=TA_CENTER)

    story: list = []
    gen = _gen_date()
    org_name = context.get("org", "") if context else ""
    tcfg = tcfg or TEMPLATES["technical"]
    avail_w = page_w - 4 * cm

    # ── Compliance: classification banner ──────────────────────────────────
    if template == "compliance":
        cls_data = [[Paragraph(
            "CONFIDENTIAL — FOR INTERNAL USE ONLY",
            ParagraphStyle("cls", fontSize=9, textColor=RED, alignment=TA_CENTER,
                           fontName="Helvetica-Bold"))]]
        cls_tbl = RLTable(cls_data, colWidths=[avail_w])
        cls_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), HexColor("#FEE2E2")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("ROUNDEDCORNERS", [4, 4, 4, 4]),
        ]))
        story.append(cls_tbl)
        story.append(Spacer(1, 3 * mm))

    # ── Header banner ──────────────────────────────────────────────────────
    header_data = [[
        Paragraph(title, s_title),
    ], [
        Paragraph(f"Organisation: {_safe_html(org_name)}  |  {gen}", s_subtitle),
    ]]
    header_table = RLTable(header_data, colWidths=[avail_w])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NAVY),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, -1), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.4 * cm))

    # ── KPI cards ──────────────────────────────────────────────────────────
    summary = context.get("summary") if context else None
    if summary:
        total_open = summary.open_dependabot + summary.open_code_scanning + summary.open_secret_scanning
        total_crit = summary.critical_dependabot + summary.critical_code_scanning
        total_high = summary.high_dependabot + summary.high_code_scanning

        kpis = [
            (str(total_open), "TOTAL OPEN", RED),
            (str(total_crit), "CRITICAL", RED),
            (str(total_high), "HIGH", ORANGE),
            (str(summary.repositories_affected), "REPOS AFFECTED", BLUE),
        ]

        kpi_cells = []
        for val, lbl, colour in kpis:
            s_v = ParagraphStyle("kv", parent=s_kpi_val, textColor=colour)
            kpi_cells.append([
                [Paragraph(val, s_v)],
                [Paragraph(lbl, s_kpi_lbl)],
            ])

        # Build a 1-row x 4-col table; each cell is a 2-row inner table (value over label)
        card_w = avail_w / 4 - 2 * mm
        kpi_inner = []
        for cell_data in kpi_cells:
            inner = RLTable(cell_data, colWidths=[card_w], rowHeights=[28, 14])
            inner.setStyle(TableStyle([
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (0, 0), "BOTTOM"),
                ("VALIGN", (0, 1), (0, 1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]))
            kpi_inner.append(inner)

        kpi_table = RLTable([kpi_inner], colWidths=[card_w + 2 * mm] * 4)
        kpi_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (0, 0), 0.5, BORDER),
            ("BOX", (1, 0), (1, 0), 0.5, BORDER),
            ("BOX", (2, 0), (2, 0), 0.5, BORDER),
            ("BOX", (3, 0), (3, 0), 0.5, BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 0.5 * cm))

        # ── Summary table ──────────────────────────────────────────────────
        sum_data = [
            ["Category", "Total", "Open", "Critical", "High"],
            ["Dependabot", str(summary.total_dependabot), str(summary.open_dependabot),
             str(summary.critical_dependabot), str(summary.high_dependabot)],
            ["Code Scanning", str(summary.total_code_scanning), str(summary.open_code_scanning),
             str(summary.critical_code_scanning), str(summary.high_code_scanning)],
            ["Secret Scanning", str(summary.total_secret_scanning), str(summary.open_secret_scanning),
             "--", "--"],
        ]
        col_widths = [avail_w * 0.3, avail_w * 0.15, avail_w * 0.15, avail_w * 0.2, avail_w * 0.2]
        sum_table = RLTable(sum_data, colWidths=col_widths, repeatRows=1)
        sum_style = [
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, LIGHT_BG]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]
        # Highlight critical/high cells with colour
        for row_idx in (1, 2):
            crit_val = int(sum_data[row_idx][3]) if sum_data[row_idx][3] != "--" else 0
            high_val = int(sum_data[row_idx][4]) if sum_data[row_idx][4] != "--" else 0
            if crit_val > 0:
                sum_style.append(("TEXTCOLOR", (3, row_idx), (3, row_idx), RED))
                sum_style.append(("FONTNAME", (3, row_idx), (3, row_idx), "Helvetica-Bold"))
            if high_val > 0:
                sum_style.append(("TEXTCOLOR", (4, row_idx), (4, row_idx), ORANGE))
                sum_style.append(("FONTNAME", (4, row_idx), (4, row_idx), "Helvetica-Bold"))

        sum_table.setStyle(TableStyle(sum_style))
        story.append(sum_table)

        if summary.push_protection_bypassed > 0:
            story.append(Spacer(1, 4 * mm))
            warn_data = [[Paragraph(
                f"<b>WARNING:</b> {summary.push_protection_bypassed} push protection bypass(es) detected!",
                ParagraphStyle("warn", parent=s_body, textColor=white, alignment=TA_CENTER,
                               fontName="Helvetica-Bold", fontSize=10))]]
            warn_tbl = RLTable(warn_data, colWidths=[avail_w])
            warn_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), RED),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("ROUNDEDCORNERS", [4, 4, 4, 4]),
            ]))
            story.append(warn_tbl)

        # Top vulnerable repos
        if summary.top_vulnerable_repos:
            story.append(Spacer(1, 0.4 * cm))
            story.append(Paragraph("Top Vulnerable Repositories", s_h3))
            for i, repo_name in enumerate(summary.top_vulnerable_repos, 1):
                story.append(Paragraph(f"{i}. {_safe_html(repo_name)}", s_num))

        story.append(Spacer(1, 0.5 * cm))
        story.append(HRFlowable(width="100%", color=BORDER, thickness=1, spaceAfter=8))

    # ── AI Analysis content ────────────────────────────────────────────────
    story.append(Paragraph("AI Analysis", s_h2))
    story.append(Spacer(1, 2 * mm))

    # Style for section headings (e.g. "1. RISK SCORE SUMMARY")
    s_section = ParagraphStyle("RptSection", parent=styles["Heading2"],
        fontSize=10, textColor=white, spaceBefore=12, spaceAfter=6,
        fontName="Helvetica-Bold", leftIndent=0)
    # Sub-heading style for bold-only lines (e.g. "**Repository Analysis:**")
    s_subhead = ParagraphStyle("RptSubHead", parent=styles["Heading3"],
        fontSize=10, textColor=NAVY, spaceBefore=10, spaceAfter=4,
        fontName="Helvetica-Bold")
    # Table cell styles
    s_tbl_header = ParagraphStyle("TblHeader", fontSize=8, leading=10,
        textColor=white, fontName="Helvetica-Bold")
    s_tbl_cell = ParagraphStyle("TblCell", fontSize=8, leading=11,
        textColor=black, fontName="Helvetica")

    def _sanitize_text(text: str) -> str:
        """Replace Unicode chars that Helvetica cannot render."""
        # Common problematic chars → safe ASCII/Latin equivalents
        replacements = {
            '\u25a0': '-', '\u25aa': '-', '\u25ab': '-',  # ■ ▪ ▫
            '\u2010': '-', '\u2011': '-', '\u2012': '-',  # various hyphens
            '\u2013': '-', '\u2014': '--', '\u2015': '--', # en-dash, em-dash
            '\u2018': "'", '\u2019': "'",                  # curly quotes
            '\u201c': '"', '\u201d': '"',
            '\u2026': '...', '\u2022': '*',                 # ellipsis, bullet
            '\u00ad': '-',                                   # soft hyphen
            '\u200b': '', '\u200c': '', '\u200d': '',       # zero-width chars
            '\ufeff': '',                                    # BOM
        }
        for orig, repl in replacements.items():
            text = text.replace(orig, repl)
        return text

    def _is_section_heading(text: str) -> bool:
        """Detect numbered section headings in any markdown variant."""
        clean = _sanitize_text(re.sub(r'\*\*', '', text)).strip()
        # Match: "1. RISK SCORE SUMMARY", "2. TOP 5 AT-RISK REPOS", etc.
        m = re.match(r'^\d+[\.\)]\s+([A-Z][A-Z0-9 :\-/&\(\),]{3,})', clean)
        if m:
            heading_part = m.group(1)
            # Require at least 2 uppercase words (2+ consecutive letters each)
            # to avoid single-word names like "MSOSE" or "DEMS"
            upper_words = re.findall(r'[A-Z]{2,}', heading_part)
            if len(upper_words) < 2:
                return False
            alpha = [c for c in heading_part if c.isalpha()]
            if alpha and sum(1 for c in alpha if c.isupper()) / len(alpha) > 0.6:
                return True
        return False

    def _is_bold_only_line(text: str) -> bool:
        """Detect lines that are entirely bold like **Some Heading**."""
        stripped = text.strip()
        return (stripped.startswith('**') and stripped.endswith('**')
                and stripped.count('**') == 2 and len(stripped) > 6)

    def _is_table_separator(text: str) -> bool:
        """Detect markdown table separator lines like |---|---|."""
        return bool(re.match(r'^\|[\s\-:|]+$', text.rstrip()))

    def _is_table_row(text: str) -> bool:
        """Detect markdown table data rows like | val | val |."""
        # Accept rows with or without trailing pipe: '| a | b |' and '| a | b'
        return text.startswith('|') and text.count('|') >= 3

    def _parse_table_row(text: str) -> list[str]:
        """Split a markdown table row into cell values."""
        # Strip leading/trailing pipes and whitespace, then split
        cells = text.strip().strip('|').split('|')
        return [c.strip() for c in cells]

    def _build_markdown_table(rows: list[list[str]]) -> RLTable:
        """Convert parsed markdown table rows into a styled reportlab Table."""
        if not rows:
            return Spacer(1, 1)
        # First row is header
        header = rows[0]
        data_rows = rows[1:]
        n_cols = len(header)

        # Build table data with Paragraphs
        tbl_data = [[Paragraph(_sanitize_text(_safe_html(c)), s_tbl_header) for c in header]]
        for row in data_rows:
            # Pad or trim row to match header column count
            padded = (row + [''] * n_cols)[:n_cols]
            tbl_data.append([Paragraph(
                re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>',
                       _sanitize_text(_safe_html(c))),
                s_tbl_cell) for c in padded])

        # Calculate column widths proportionally
        col_widths = [avail_w / n_cols] * n_cols

        tbl = RLTable(tbl_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), white),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, LIGHT_BG]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        return tbl

    def _make_section_banner(text: str) -> RLTable:
        clean = re.sub(r'\*\*', '', text).strip()
        sec_data = [[Paragraph(_sanitize_text(_safe_html(clean)), s_section)]]
        sec_tbl = RLTable(sec_data, colWidths=[avail_w])
        sec_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), DARK_BLUE),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("ROUNDEDCORNERS", [4, 4, 4, 4]),
        ]))
        return sec_tbl

    def _fmt_inline(text: str) -> str:
        """Convert **bold** markers, sanitize chars, and escape HTML."""
        return re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", _sanitize_text(_safe_html(text)))

    # ── Parse content lines, accumulating table rows ──────────────────────
    lines = content.split("\n")
    pending_table_rows: list[list[str]] = []

    def _flush_table():
        """Flush any accumulated table rows into a rendered table."""
        if pending_table_rows:
            story.append(Spacer(1, 2 * mm))
            story.append(_build_markdown_table(pending_table_rows))
            story.append(Spacer(1, 2 * mm))
            pending_table_rows.clear()

    for line in lines:
        stripped = line.strip()

        # ── Markdown table handling ────────────────────────────────────
        if _is_table_separator(stripped):
            continue  # skip separator lines like |---|---|
        if _is_table_row(stripped):
            pending_table_rows.append(_parse_table_row(stripped))
            continue
        # If we had been accumulating table rows and hit a non-table line, flush
        _flush_table()

        if not stripped:
            story.append(Spacer(1, 2 * mm))
        # Numbered section headings: **1. RISK SCORE SUMMARY** etc.
        elif _is_section_heading(stripped):
            story.append(_make_section_banner(stripped))
        # Markdown headings
        elif stripped.startswith("### "):
            story.append(Paragraph(_sanitize_text(_safe_html(stripped[4:])), s_h3))
        elif stripped.startswith("## "):
            story.append(Paragraph(_sanitize_text(_safe_html(stripped[3:])), s_h2))
        elif stripped.startswith("# "):
            story.append(Paragraph(_sanitize_text(_safe_html(stripped[2:])), s_h2))
        # Separator lines
        elif re.match(r"^[=\-]{3,}$", stripped):
            story.append(HRFlowable(width="100%", color=BORDER, thickness=0.5, spaceAfter=4, spaceBefore=4))
        # ALL-CAPS standalone lines (e.g. "RISK ASSESSMENT")
        elif re.sub(r'\*\*', '', stripped).strip().isupper() and len(re.sub(r'\*\*', '', stripped).strip()) > 5:
            story.append(_make_section_banner(stripped))
        # Bold-only lines as sub-headings (e.g. "**Repository Analysis:**")
        elif _is_bold_only_line(stripped):
            heading_text = stripped.strip('* ')
            story.append(Paragraph(_sanitize_text(_safe_html(heading_text)), s_subhead))
        # Numbered items (e.g. "1. Upgrade lodash...")
        elif re.match(r"^\d+[\.\)]\s", stripped):
            story.append(Paragraph(_fmt_inline(stripped), s_num))
        # Bullet items
        elif stripped.startswith("- ") or stripped.startswith("* "):
            story.append(Paragraph(_fmt_inline(stripped[2:]), s_bullet, bulletText="\u2022"))
        # Body text
        else:
            story.append(Paragraph(_fmt_inline(stripped), s_body))

    # Flush any remaining table rows at end of content
    _flush_table()

    # ── Footer ─────────────────────────────────────────────────────────────
    story.append(Spacer(1, 1 * cm))

    # Compliance: sign-off block
    if template == "compliance":
        story.append(HRFlowable(width="100%", color=BORDER, thickness=1, spaceAfter=12))
        signoff_data = [
            ["Prepared by:", "AI Git Guard — Automated Security Analysis"],
            ["Date:", gen],
            ["Classification:", "CONFIDENTIAL — FOR INTERNAL USE ONLY"],
            ["Reviewed by:", "____________________________"],
            ["Approved by:", "____________________________"],
        ]
        signoff_tbl = RLTable(signoff_data, colWidths=[avail_w * 0.25, avail_w * 0.75])
        signoff_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TEXTCOLOR", (0, 0), (-1, -1), GREY),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, -1), (-1, -1), 0.5, BORDER),
        ]))
        story.append(signoff_tbl)
        story.append(Spacer(1, 0.5 * cm))

    story.append(HRFlowable(width="100%", color=BORDER, thickness=0.5, spaceAfter=6))
    tpl_label = TEMPLATES.get(template, {}).get("label", "")
    story.append(Paragraph(f"AI Git Guard -- GitHub Advanced Security AI Agent  |  Template: {tpl_label}", s_footer))

    doc.build(story)
    logger.info("PDF report written: %s", path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
#  HTML
# ══════════════════════════════════════════════════════════════════════════════

def _render_html(content: str, title: str, path: Path, context: Optional[dict[str, Any]] = None,
                 template: ReportTemplate = "technical", tcfg: dict | None = None) -> Path:
    org_name = _safe_html(context.get("org", "")) if context else ""
    gen = _gen_date()
    summary = context.get("summary") if context else None
    tcfg = tcfg or TEMPLATES["technical"]

    # ── Metrics cards + summary table ──────────────────────────────────────
    summary_html = ""
    alert_tables_html = ""
    if summary:
        total_open = summary.open_dependabot + summary.open_code_scanning + summary.open_secret_scanning
        total_crit = summary.critical_dependabot + summary.critical_code_scanning
        total_high = summary.high_dependabot + summary.high_code_scanning

        push_warn = ""
        if summary.push_protection_bypassed > 0:
            push_warn = f'<div class="push-warn">WARNING: {summary.push_protection_bypassed} push protection bypass(es) detected!</div>'

        top_repos_html = ""
        if summary.top_vulnerable_repos:
            repos_li = "".join(f"<li>{_safe_html(r)}</li>" for r in summary.top_vulnerable_repos)
            top_repos_html = f'<div class="card" style="margin-top:16px"><h3>Top Vulnerable Repositories</h3><ol class="repo-list">{repos_li}</ol></div>'

        summary_html = f"""
  <div class="kpi-grid">
    <div class="kpi-card"><div class="kpi-value red">{total_open}</div><div class="kpi-label">Total Open Alerts</div></div>
    <div class="kpi-card"><div class="kpi-value red">{total_crit}</div><div class="kpi-label">Critical</div></div>
    <div class="kpi-card"><div class="kpi-value orange">{total_high}</div><div class="kpi-label">High</div></div>
    <div class="kpi-card"><div class="kpi-value blue">{summary.repositories_affected}</div><div class="kpi-label">Repos Affected</div></div>
  </div>
  {push_warn}
  <div class="card">
    <table class="data-table summary-table">
      <thead><tr><th>Category</th><th>Total</th><th>Open</th><th>Critical</th><th>High</th></tr></thead>
      <tbody>
        <tr><td>Dependabot</td><td>{summary.total_dependabot}</td><td>{summary.open_dependabot}</td>
            <td class="{"crit" if summary.critical_dependabot > 0 else ""}">{summary.critical_dependabot}</td>
            <td class="{"high" if summary.high_dependabot > 0 else ""}">{summary.high_dependabot}</td></tr>
        <tr><td>Code Scanning</td><td>{summary.total_code_scanning}</td><td>{summary.open_code_scanning}</td>
            <td class="{"crit" if summary.critical_code_scanning > 0 else ""}">{summary.critical_code_scanning}</td>
            <td class="{"high" if summary.high_code_scanning > 0 else ""}">{summary.high_code_scanning}</td></tr>
        <tr><td>Secret Scanning</td><td>{summary.total_secret_scanning}</td><td>{summary.open_secret_scanning}</td>
            <td colspan="2" style="text-align:center;color:#94a3b8">--</td></tr>
      </tbody>
    </table>
  </div>
  {top_repos_html}"""

    # ── Alert detail tables ────────────────────────────────────────────────
    dep_alerts = context.get("dependabot", []) if context else []
    cs_alerts = context.get("code_scanning", []) if context else []
    ss_alerts = context.get("secret_scanning", []) if context else []

    sev_class = lambda s: f'sev-{s.value.lower()}' if hasattr(s, 'value') else 'sev-unknown'

    if dep_alerts:
        dep_rows = ""
        for a in dep_alerts[:50]:
            sev = a.severity.value.upper()
            sc = a.severity.value.lower()
            dep_rows += f"""<tr>
              <td>{a.alert_number}</td><td>{_safe_html(a.repository.full_name)}</td>
              <td><strong>{_safe_html(a.package.name)}</strong></td><td>{_safe_html(a.package.ecosystem)}</td>
              <td><span class="badge {sc}">{sev}</span></td>
              <td>{_safe_html(a.advisory.cve_id or a.advisory.ghsa_id or "")}</td>
              <td>{_safe_html((a.advisory.summary or "")[:100])}</td>
              <td>{a.state.value.upper()}</td>
              <td>{_safe_html(a.patched_version or "N/A")}</td>
              <td>{a.created_at.strftime("%Y-%m-%d") if a.created_at else ""}</td></tr>"""
        alert_tables_html += f"""
  <div class="card">
    <h3>Dependabot Alerts <span class="count">({len(dep_alerts)})</span></h3>
    <div class="table-scroll">
    <table class="data-table"><thead><tr>
      <th>#</th><th>Repository</th><th>Package</th><th>Ecosystem</th><th>Severity</th>
      <th>CVE / GHSA</th><th>Summary</th><th>State</th><th>Patched</th><th>Created</th>
    </tr></thead><tbody>{dep_rows}</tbody></table>
    </div>
    {"<p class='more'>Showing first 50 of " + str(len(dep_alerts)) + " alerts</p>" if len(dep_alerts) > 50 else ""}
  </div>"""

    if cs_alerts:
        cs_rows = ""
        for a in cs_alerts[:50]:
            sev = a.severity.value.upper()
            sc = a.severity.value.lower()
            loc = f"{a.location.path}:{a.location.start_line}" if a.location and a.location.path else ""
            cs_rows += f"""<tr>
              <td>{a.alert_number}</td><td>{_safe_html(a.repository.full_name)}</td>
              <td><strong>{_safe_html(a.rule.name or a.rule.id)}</strong></td>
              <td><span class="badge {sc}">{sev}</span></td>
              <td>{_safe_html(a.tool_name or "")}</td>
              <td class="mono">{_safe_html(loc)}</td>
              <td>{_safe_html((a.message or "")[:100])}</td>
              <td>{a.state.value.upper()}</td>
              <td>{a.created_at.strftime("%Y-%m-%d") if a.created_at else ""}</td></tr>"""
        alert_tables_html += f"""
  <div class="card">
    <h3>Code Scanning Alerts <span class="count">({len(cs_alerts)})</span></h3>
    <div class="table-scroll">
    <table class="data-table"><thead><tr>
      <th>#</th><th>Repository</th><th>Rule</th><th>Severity</th><th>Tool</th>
      <th>Location</th><th>Message</th><th>State</th><th>Created</th>
    </tr></thead><tbody>{cs_rows}</tbody></table>
    </div>
    {"<p class='more'>Showing first 50 of " + str(len(cs_alerts)) + " alerts</p>" if len(cs_alerts) > 50 else ""}
  </div>"""

    if ss_alerts:
        ss_rows = ""
        for a in ss_alerts[:50]:
            bypassed = '<span class="badge critical">YES</span>' if a.push_protection_bypassed else "No"
            ss_rows += f"""<tr>
              <td>{a.alert_number}</td><td>{_safe_html(a.repository.full_name)}</td>
              <td>{_safe_html(a.secret_type_display_name or a.secret_type or "")}</td>
              <td>{a.state.value.upper()}</td><td>{bypassed}</td>
              <td>{a.created_at.strftime("%Y-%m-%d") if a.created_at else ""}</td></tr>"""
        alert_tables_html += f"""
  <div class="card">
    <h3>Secret Scanning Alerts <span class="count">({len(ss_alerts)})</span></h3>
    <div class="table-scroll">
    <table class="data-table"><thead><tr>
      <th>#</th><th>Repository</th><th>Secret Type</th><th>State</th><th>Push Bypassed</th><th>Created</th>
    </tr></thead><tbody>{ss_rows}</tbody></table>
    </div>
  </div>"""

    # ── AI Analysis content ────────────────────────────────────────────────
    content_html = _content_to_html(content)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{_safe_html(title)}</title>
  <style>
    :root {{
      --navy: #1a1a2e; --dark-blue: #16213e; --red: #e63946; --orange: #f4845f;
      --amber: #f59e0b; --green: #10b981; --blue: #3b82f6; --purple: #7c3aed;
      --bg: #f0f2f5; --card: #ffffff; --border: #e2e8f0; --text: #1e293b;
      --muted: #64748b; --light: #f8fafc;
    }}
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ font-family: 'Segoe UI', system-ui, -apple-system, sans-serif; background: var(--bg); color: var(--text); line-height: 1.6; }}
    .page {{ max-width: 1140px; margin: 0 auto; padding: 32px 24px; }}

    /* Header */
    .header {{ background: linear-gradient(135deg, var(--navy) 0%, var(--dark-blue) 60%, #0f3460 100%); color: #fff; padding: 36px 44px 28px; border-radius: 14px; margin-bottom: 28px; position: relative; overflow: hidden; }}
    .header::after {{ content: ''; position: absolute; top: -50%; right: -10%; width: 300px; height: 300px; background: rgba(255,255,255,0.03); border-radius: 50%; }}
    .header h1 {{ font-size: 1.75em; font-weight: 700; letter-spacing: -0.5px; margin-bottom: 6px; }}
    .header .meta {{ color: #94a3b8; font-size: 0.82em; letter-spacing: 0.2px; }}
    .header .meta span {{ color: #cbd5e1; font-weight: 600; }}

    /* KPI Cards */
    .kpi-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 20px; }}
    .kpi-card {{ background: var(--card); border-radius: 12px; padding: 22px 16px 18px; text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,0.06); border: 1px solid var(--border); transition: transform 0.15s, box-shadow 0.15s; }}
    .kpi-card:hover {{ transform: translateY(-2px); box-shadow: 0 4px 12px rgba(0,0,0,0.1); }}
    .kpi-value {{ font-size: 2.4em; font-weight: 800; line-height: 1.1; }}
    .kpi-value.red {{ color: var(--red); }}
    .kpi-value.orange {{ color: var(--orange); }}
    .kpi-value.blue {{ color: var(--blue); }}
    .kpi-label {{ font-size: 0.72em; color: var(--muted); text-transform: uppercase; letter-spacing: 0.8px; margin-top: 6px; font-weight: 600; }}

    /* Cards */
    .card {{ background: var(--card); border-radius: 12px; padding: 28px 32px; box-shadow: 0 1px 4px rgba(0,0,0,0.06); border: 1px solid var(--border); margin-bottom: 20px; }}
    .card h2 {{ color: var(--navy); font-size: 1.2em; margin-bottom: 16px; padding-bottom: 10px; border-bottom: 2px solid var(--border); }}
    .card h3 {{ color: var(--navy); font-size: 1.05em; margin-bottom: 14px; }}
    .card h3 .count {{ color: var(--muted); font-weight: 400; font-size: 0.85em; }}

    /* Push protection warning */
    .push-warn {{ background: var(--red); color: #fff; padding: 12px 20px; border-radius: 8px; font-weight: 700; text-align: center; margin-bottom: 20px; font-size: 0.95em; }}

    /* Tables */
    .table-scroll {{ overflow-x: auto; margin: 0 -8px; padding: 0 8px; }}
    .data-table {{ width: 100%; border-collapse: collapse; font-size: 0.85em; }}
    .data-table th {{ background: var(--navy); color: #fff; padding: 10px 14px; text-align: left; font-size: 0.8em; text-transform: uppercase; letter-spacing: 0.5px; white-space: nowrap; font-weight: 600; }}
    .data-table th:first-child {{ border-radius: 8px 0 0 0; }}
    .data-table th:last-child {{ border-radius: 0 8px 0 0; }}
    .data-table td {{ padding: 9px 14px; border-bottom: 1px solid var(--border); vertical-align: top; }}
    .data-table tbody tr:nth-child(even) {{ background: var(--light); }}
    .data-table tbody tr:hover {{ background: #eef2ff; }}
    .data-table .mono {{ font-family: 'Cascadia Code', 'Fira Code', monospace; font-size: 0.88em; color: var(--muted); }}
    .data-table td.crit {{ color: var(--red); font-weight: 700; }}
    .data-table td.high {{ color: var(--orange); font-weight: 700; }}
    .summary-table {{ margin: 0; }}
    .summary-table td:first-child {{ font-weight: 600; }}
    p.more {{ color: var(--muted); font-size: 0.8em; margin-top: 10px; font-style: italic; }}

    /* Severity badges */
    .badge {{ display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 0.78em; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px; }}
    .badge.critical {{ background: #fee2e2; color: var(--red); }}
    .badge.high {{ background: #ffedd5; color: #c2410c; }}
    .badge.medium {{ background: #fef3c7; color: #b45309; }}
    .badge.low {{ background: #d1fae5; color: #047857; }}
    .badge.warning {{ background: #fef3c7; color: #b45309; }}
    .badge.note {{ background: #dbeafe; color: #1d4ed8; }}
    .badge.none, .badge.unknown {{ background: #f1f5f9; color: var(--muted); }}

    /* Repo list */
    .repo-list {{ padding-left: 20px; }}
    .repo-list li {{ padding: 4px 0; font-family: 'Cascadia Code', 'Fira Code', monospace; font-size: 0.92em; color: var(--dark-blue); }}

    /* AI Analysis content styling */
    .analysis {{ line-height: 1.8; }}
    .analysis p {{ margin-bottom: 10px; }}
    .analysis h2.section-h2 {{ color: var(--navy); font-size: 1.15em; margin: 20px 0 8px; padding-bottom: 6px; border-bottom: 2px solid var(--border); }}
    .analysis h3.section-h3 {{ color: var(--dark-blue); font-size: 1.05em; margin: 16px 0 6px; }}
    .analysis h4.section-h4 {{ color: var(--text); font-size: 0.95em; margin: 12px 0 4px; }}
    .analysis hr.section-rule {{ border: none; border-top: 1px solid var(--border); margin: 16px 0; }}
    .analysis .numbered-item {{ padding: 6px 0 6px 6px; border-left: 3px solid var(--blue); margin: 4px 0 4px 8px; padding-left: 12px; }}
    .analysis ul.bullet-list {{ padding-left: 24px; margin: 6px 0; }}
    .analysis ul.bullet-list li {{ padding: 3px 0; }}
    .analysis strong {{ color: var(--navy); }}

    /* Footer */
    footer {{ text-align: center; color: var(--muted); font-size: 0.78em; padding: 24px 0 8px; letter-spacing: 0.3px; }}

    @media (max-width: 768px) {{
      .kpi-grid {{ grid-template-columns: repeat(2, 1fr); }}
      .header {{ padding: 24px; }}
      .card {{ padding: 20px; }}
    }}
    @media print {{
      body {{ background: #fff; }}
      .card {{ box-shadow: none; border: 1px solid #ddd; break-inside: avoid; }}
      .kpi-card {{ box-shadow: none; border: 1px solid #ddd; }}
    }}

    /* Template: Executive — larger KPIs, softer palette */
    body.tpl-executive .kpi-value {{ font-size: 3em; }}
    body.tpl-executive .kpi-card {{ padding: 28px 20px 22px; }}
    body.tpl-executive .header {{ background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 100%); }}

    /* Template: Compliance — formal header banner */
    .compliance-banner {{ background: #fee2e2; color: var(--red); text-align: center; padding: 8px 16px;
      font-weight: 700; font-size: 0.82em; letter-spacing: 1px; border-radius: 6px; margin-bottom: 16px;
      border: 1px solid #fca5a5; text-transform: uppercase; }}
    .signoff {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px;
      padding: 28px 32px; margin-top: 20px; }}
    .signoff h3 {{ color: var(--navy); margin-bottom: 16px; font-size: 1em; }}
    .signoff-grid {{ display: grid; grid-template-columns: 160px 1fr; gap: 12px 16px; font-size: 0.9em; }}
    .signoff-grid .label {{ color: var(--muted); font-weight: 600; }}
    .signoff-grid .value {{ color: var(--text); border-bottom: 1px solid var(--border); padding-bottom: 4px; }}
    .tpl-badge {{ display: inline-block; background: var(--light); color: var(--muted); padding: 2px 10px;
      border-radius: 12px; font-size: 0.72em; font-weight: 600; letter-spacing: 0.5px; margin-left: 8px; }}
  </style>
</head>
<body class="tpl-{template}">
  <div class="page">
    {"<div class='compliance-banner'>CONFIDENTIAL — FOR INTERNAL USE ONLY</div>" if template == "compliance" else ""}
    <div class="header">
      <h1>{_safe_html(title)} <span class="tpl-badge">{_safe_html(tcfg.get("label", ""))}</span></h1>
      <p class="meta">Generated by <span>AI Git Guard</span> &nbsp;|&nbsp; {gen}{f" &nbsp;|&nbsp; Organisation: <span>{org_name}</span>" if org_name else ""}</p>
    </div>
    {summary_html}
    <div class="card">
      <h2>AI Analysis</h2>
      <div class="analysis">
        {content_html}
      </div>
    </div>
    {alert_tables_html if tcfg.get("show_alert_tables") else ""}
    {_compliance_signoff_html(gen) if template == "compliance" else ""}
    <footer>AI Git Guard -- GitHub Advanced Security AI Agent &nbsp;|&nbsp; Template: {_safe_html(tcfg.get("label", ""))}</footer>
  </div>
</body>
</html>"""

    path.write_text(html, encoding="utf-8")
    logger.info("HTML report written: %s", path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
#  Excel
# ══════════════════════════════════════════════════════════════════════════════

# --- Colour palette (matching HTML/PDF) ---
_NAVY        = "1A1A2E"
_DARK_BLUE   = "16213E"
_WHITE       = "FFFFFF"
_LIGHT_GREY  = "F0F2F5"
_LIGHT_BG    = "F8FAFC"
_BORDER_GREY = "D1D5DB"
_RED         = "E63946"
_ORANGE      = "F4845F"
_AMBER       = "F59E0B"
_GREEN       = "10B981"
_BLUE        = "3B82F6"
_DARK_TEXT   = "1E293B"
_GREY_TEXT   = "64748B"

_SEVERITY_BG = {
    "critical": "FEE2E2", "high": "FFEDD5", "medium": "FEF3C7",
    "low": "D1FAE5", "warning": "FEF3C7", "note": "DBEAFE",
    "none": _LIGHT_GREY, "unknown": _LIGHT_GREY,
}
_SEVERITY_FG = {
    "critical": _RED, "high": "C2410C", "medium": "B45309",
    "low": "047857", "warning": "B45309", "note": "1D4ED8",
    "none": _GREY_TEXT, "unknown": _GREY_TEXT,
}


def _render_excel(content: str, title: str, path: Path, context: Optional[dict[str, Any]] = None,
                  template: ReportTemplate = "technical", tcfg: dict | None = None) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.series import DataPoint
    except ImportError:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    gen = _gen_date()
    org_name = context.get("org", "") if context else ""
    summary = context.get("summary") if context else None
    tcfg = tcfg or TEMPLATES["technical"]
    show_alert_sheets = tcfg.get("excel_alert_sheets", True)

    # ── Reusable styles ────────────────────────────────────────────────────
    thin_border = Border(
        left=Side(style="thin", color=_BORDER_GREY), right=Side(style="thin", color=_BORDER_GREY),
        top=Side(style="thin", color=_BORDER_GREY), bottom=Side(style="thin", color=_BORDER_GREY),
    )
    bottom_border = Border(bottom=Side(style="thin", color=_BORDER_GREY))
    hdr_fill = PatternFill("solid", fgColor=_NAVY)
    hdr_font = Font(bold=True, size=10, color=_WHITE, name="Calibri")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(size=10, color=_DARK_TEXT, name="Calibri")
    body_align = Alignment(vertical="center", wrap_text=True)
    stripe_a = PatternFill("solid", fgColor=_WHITE)
    stripe_b = PatternFill("solid", fgColor=_LIGHT_BG)

    def _apply_header_row(ws, ncols: int, row: int = 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = hdr_align; cell.border = thin_border
        ws.row_dimensions[row].height = 30

    def _apply_data_cell(cell, row_idx: int):
        cell.font = body_font; cell.alignment = body_align; cell.border = thin_border
        cell.fill = stripe_b if row_idx % 2 == 0 else stripe_a

    def _apply_severity(cell, severity_str: str):
        sev = (severity_str or "unknown").lower()
        cell.fill = PatternFill("solid", fgColor=_SEVERITY_BG.get(sev, _LIGHT_GREY))
        cell.font = Font(bold=True, size=10, color=_SEVERITY_FG.get(sev, _DARK_TEXT), name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.value = sev.upper()

    def _set_col_widths(ws, widths: dict[str, float]):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    def _title_block(ws, text: str, sub_text: str, ncols: int):
        """Create a styled title + subtitle banner spanning ncols."""
        end_col = get_column_letter(ncols)
        ws.merge_cells(f"A1:{end_col}1")
        c1 = ws["A1"]
        c1.value = text
        c1.font = Font(bold=True, size=16, color=_WHITE, name="Calibri")
        c1.fill = PatternFill("solid", fgColor=_NAVY)
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 44

        ws.merge_cells(f"A2:{end_col}2")
        c2 = ws["A2"]
        c2.value = sub_text
        c2.font = Font(italic=True, size=9, color=_WHITE, name="Calibri")
        c2.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 22

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 1 — Dashboard
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = _NAVY
    ncols = 8
    _title_block(ws, title, f"Organisation: {org_name}   |   {gen}", ncols)
    ws.row_dimensions[3].height = 8  # spacer

    if summary:
        total_open = summary.open_dependabot + summary.open_code_scanning + summary.open_secret_scanning
        total_crit = summary.critical_dependabot + summary.critical_code_scanning
        total_high = summary.high_dependabot + summary.high_code_scanning

        # KPI row — 4 cards across 8 columns (2 cols each)
        kpis = [
            ("TOTAL OPEN ALERTS", total_open, _RED),
            ("CRITICAL", total_crit, _RED),
            ("HIGH", total_high, _ORANGE),
            ("REPOS AFFECTED", summary.repositories_affected, _BLUE),
        ]
        for j, (label, value, colour) in enumerate(kpis):
            col = j * 2 + 1

            # Card background
            for dc in (col, col + 1):
                for dr in (4, 5):
                    c = ws.cell(row=dr, column=dc)
                    c.fill = PatternFill("solid", fgColor=_WHITE)
                    c.border = Border(
                        top=Side(style="medium", color=colour) if dr == 4 else Side(style="thin", color=_BORDER_GREY),
                        bottom=Side(style="thin", color=_BORDER_GREY) if dr == 4 else Side(style="medium", color=_BORDER_GREY),
                        left=Side(style="thin", color=_BORDER_GREY) if dc == col else Side(style=None),
                        right=Side(style="thin", color=_BORDER_GREY) if dc == col + 1 else Side(style=None),
                    )

            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
            val_cell = ws.cell(row=4, column=col, value=value)
            val_cell.font = Font(bold=True, size=26, color=colour, name="Calibri")
            val_cell.alignment = Alignment(horizontal="center", vertical="center")
            val_cell.fill = PatternFill("solid", fgColor=_WHITE)
            val_cell.border = Border(top=Side(style="medium", color=colour),
                                     left=Side(style="thin", color=_BORDER_GREY),
                                     right=Side(style="thin", color=_BORDER_GREY))

            ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
            lbl_cell = ws.cell(row=5, column=col, value=label)
            lbl_cell.font = Font(size=8, color=_GREY_TEXT, name="Calibri", bold=True)
            lbl_cell.alignment = Alignment(horizontal="center", vertical="top")
            lbl_cell.fill = PatternFill("solid", fgColor=_WHITE)
            lbl_cell.border = Border(bottom=Side(style="thin", color=_BORDER_GREY),
                                     left=Side(style="thin", color=_BORDER_GREY),
                                     right=Side(style="thin", color=_BORDER_GREY))

        ws.row_dimensions[4].height = 48
        ws.row_dimensions[5].height = 20
        ws.row_dimensions[6].height = 12  # spacer

        # ── Summary table ──────────────────────────────────────────────────
        s_row = 7
        sum_headers = ["Category", "Total", "Open", "Critical", "High"]
        for c, h in enumerate(sum_headers, 1):
            ws.cell(row=s_row, column=c, value=h)
        _apply_header_row(ws, len(sum_headers), s_row)

        sum_rows = [
            ("Dependabot", summary.total_dependabot, summary.open_dependabot, summary.critical_dependabot, summary.high_dependabot),
            ("Code Scanning", summary.total_code_scanning, summary.open_code_scanning, summary.critical_code_scanning, summary.high_code_scanning),
            ("Secret Scanning", summary.total_secret_scanning, summary.open_secret_scanning, "--", "--"),
        ]
        for i, (cat, total, opn, crit, high) in enumerate(sum_rows):
            r = s_row + 1 + i
            for c, val in enumerate([cat, total, opn, crit, high], 1):
                cell = ws.cell(row=r, column=c, value=val)
                _apply_data_cell(cell, i)
                if c >= 2:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if c == 4 and isinstance(val, int) and val > 0:
                    cell.font = Font(bold=True, size=11, color=_RED, name="Calibri")
                elif c == 5 and isinstance(val, int) and val > 0:
                    cell.font = Font(bold=True, size=11, color="C2410C", name="Calibri")
            ws.row_dimensions[r].height = 24

        # ── Pie chart — Severity Distribution ──────────────────────────────
        chart_row = s_row + len(sum_rows) + 2
        ws.cell(row=chart_row, column=1, value="Severity Distribution (Open Alerts)").font = Font(
            bold=True, size=12, color=_DARK_TEXT, name="Calibri")

        data_row = chart_row + 1
        sev_data = [("Critical", total_crit), ("High", total_high),
                     ("Medium + Low", max(0, total_open - total_crit - total_high))]
        for i, (lbl, val) in enumerate(sev_data):
            ws.cell(row=data_row + i, column=1, value=lbl).font = Font(size=9, color=_GREY_TEXT, name="Calibri")
            ws.cell(row=data_row + i, column=2, value=val).font = Font(size=9, color=_DARK_TEXT, name="Calibri")

        if any(v > 0 for _, v in sev_data):
            pie = PieChart()
            pie.title = None; pie.style = 26; pie.width = 12; pie.height = 9
            labels = Reference(ws, min_col=1, min_row=data_row, max_row=data_row + 2)
            data   = Reference(ws, min_col=2, min_row=data_row, max_row=data_row + 2)
            pie.add_data(data, titles_from_data=False)
            pie.set_categories(labels)
            for idx, colour in enumerate([_RED, _ORANGE, _AMBER]):
                pt = DataPoint(idx=idx)
                pt.graphicalProperties.solidFill = colour
                pie.series[0].data_points.append(pt)
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True; pie.dataLabels.showVal = True
            ws.add_chart(pie, f"D{chart_row}")

        # ── Bar chart — Alerts by Category ─────────────────────────────────
        bar_row = chart_row
        bar_data_row = chart_row + 1
        ws.cell(row=bar_data_row, column=6, value="Dependabot").font = Font(size=9, name="Calibri")
        ws.cell(row=bar_data_row, column=7, value=summary.open_dependabot).font = Font(size=9, name="Calibri")
        ws.cell(row=bar_data_row + 1, column=6, value="Code Scanning").font = Font(size=9, name="Calibri")
        ws.cell(row=bar_data_row + 1, column=7, value=summary.open_code_scanning).font = Font(size=9, name="Calibri")
        ws.cell(row=bar_data_row + 2, column=6, value="Secret Scanning").font = Font(size=9, name="Calibri")
        ws.cell(row=bar_data_row + 2, column=7, value=summary.open_secret_scanning).font = Font(size=9, name="Calibri")

        bar = BarChart()
        bar.type = "col"; bar.style = 26; bar.width = 12; bar.height = 9
        bar.title = "Open Alerts by Category"
        bar.y_axis.title = "Count"
        bar_labels = Reference(ws, min_col=6, min_row=bar_data_row, max_row=bar_data_row + 2)
        bar_data   = Reference(ws, min_col=7, min_row=bar_data_row, max_row=bar_data_row + 2)
        bar.add_data(bar_data, titles_from_data=False)
        bar.set_categories(bar_labels)
        bar.series[0].graphicalProperties.solidFill = _BLUE
        bar.legend = None
        ws.add_chart(bar, f"D{chart_row + 10}")

        # ── Top vulnerable repos ───────────────────────────────────────────
        repos_row = chart_row + 21
        if summary.top_vulnerable_repos:
            ws.merge_cells(f"A{repos_row}:C{repos_row}")
            ws.cell(row=repos_row, column=1, value="Top Vulnerable Repositories").font = Font(
                bold=True, size=12, color=_DARK_TEXT, name="Calibri")
            for c, h in enumerate(["#", "Repository"], 1):
                ws.cell(row=repos_row + 1, column=c, value=h)
            _apply_header_row(ws, 2, repos_row + 1)
            for i, rname in enumerate(summary.top_vulnerable_repos):
                r = repos_row + 2 + i
                cell_num = ws.cell(row=r, column=1, value=i + 1)
                cell_name = ws.cell(row=r, column=2, value=rname)
                _apply_data_cell(cell_num, i)
                _apply_data_cell(cell_name, i)
                cell_num.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[r].height = 22

        # Push protection warning
        if summary.push_protection_bypassed > 0:
            wr = repos_row + len(summary.top_vulnerable_repos) + 4 if summary.top_vulnerable_repos else repos_row + 2
            end_col = get_column_letter(ncols)
            ws.merge_cells(f"A{wr}:{end_col}{wr}")
            warn = ws.cell(row=wr, column=1,
                           value=f"WARNING: {summary.push_protection_bypassed} push protection bypass(es) detected!")
            warn.font = Font(bold=True, size=11, color=_WHITE, name="Calibri")
            warn.fill = PatternFill("solid", fgColor=_RED)
            warn.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[wr].height = 32

    _set_col_widths(ws, {"A": 22, "B": 14, "C": 14, "D": 16, "E": 14, "F": 18, "G": 14, "H": 14})

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 2 — Dependabot Alerts
    # ══════════════════════════════════════════════════════════════════════
    dep_alerts = context.get("dependabot", []) if context else []
    if dep_alerts and show_alert_sheets:
        ws_d = wb.create_sheet("Dependabot Alerts")
        ws_d.sheet_properties.tabColor = _RED
        dep_cols = ["#", "Repository", "Package", "Ecosystem", "Severity",
                    "CVE / GHSA", "Summary", "State", "Vuln Range", "Patched", "Created", "URL"]
        ncols_d = len(dep_cols)
        _title_block(ws_d, "Dependabot Vulnerability Alerts", f"{len(dep_alerts)} alerts  |  {org_name}  |  {gen}", ncols_d)
        ws_d.row_dimensions[3].height = 6

        for c, h in enumerate(dep_cols, 1):
            ws_d.cell(row=4, column=c, value=h)
        _apply_header_row(ws_d, ncols_d, 4)
        end_letter = get_column_letter(ncols_d)
        ws_d.auto_filter.ref = f"A4:{end_letter}4"
        ws_d.freeze_panes = "A5"

        for i, a in enumerate(dep_alerts):
            r = i + 5
            vals = [
                a.alert_number, a.repository.full_name, a.package.name, a.package.ecosystem,
                None, a.advisory.cve_id or a.advisory.ghsa_id or "",
                (a.advisory.summary or "")[:120], a.state.value.upper(),
                a.vulnerable_version_range or "", a.patched_version or "N/A",
                a.created_at.strftime("%Y-%m-%d") if a.created_at else "", a.html_url or "",
            ]
            for c, val in enumerate(vals, 1):
                cell = ws_d.cell(row=r, column=c, value=val)
                if c == 5:
                    _apply_severity(cell, a.severity.value)
                else:
                    _apply_data_cell(cell, i)
                    if c == 1:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_d.row_dimensions[r].height = 22

        _set_col_widths(ws_d, {
            "A": 6, "B": 28, "C": 22, "D": 12, "E": 12, "F": 18,
            "G": 50, "H": 10, "I": 16, "J": 14, "K": 12, "L": 40,
        })

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 3 — Code Scanning Alerts
    # ══════════════════════════════════════════════════════════════════════
    cs_alerts = context.get("code_scanning", []) if context else []
    if cs_alerts and show_alert_sheets:
        ws_c = wb.create_sheet("Code Scanning Alerts")
        ws_c.sheet_properties.tabColor = _ORANGE
        cs_cols = ["#", "Repository", "Rule", "Severity", "Tool", "File", "Line", "Message", "State", "Created", "URL"]
        ncols_c = len(cs_cols)
        _title_block(ws_c, "Code Scanning Alerts (SAST)", f"{len(cs_alerts)} alerts  |  {org_name}  |  {gen}", ncols_c)
        ws_c.row_dimensions[3].height = 6

        for c, h in enumerate(cs_cols, 1):
            ws_c.cell(row=4, column=c, value=h)
        _apply_header_row(ws_c, ncols_c, 4)
        ws_c.auto_filter.ref = f"A4:{get_column_letter(ncols_c)}4"
        ws_c.freeze_panes = "A5"

        for i, a in enumerate(cs_alerts):
            r = i + 5
            loc_path = a.location.path if a.location else ""
            loc_line = a.location.start_line if a.location else ""
            vals = [
                a.alert_number, a.repository.full_name, a.rule.name or a.rule.id,
                None, a.tool_name or "", loc_path, loc_line,
                (a.message or "")[:120], a.state.value.upper(),
                a.created_at.strftime("%Y-%m-%d") if a.created_at else "", a.html_url or "",
            ]
            for c, val in enumerate(vals, 1):
                cell = ws_c.cell(row=r, column=c, value=val)
                if c == 4:
                    _apply_severity(cell, a.severity.value)
                else:
                    _apply_data_cell(cell, i)
                    if c in (1, 7):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_c.row_dimensions[r].height = 22

        _set_col_widths(ws_c, {
            "A": 6, "B": 28, "C": 30, "D": 12, "E": 14, "F": 36,
            "G": 8, "H": 50, "I": 10, "J": 12, "K": 40,
        })

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 4 — Secret Scanning Alerts
    # ══════════════════════════════════════════════════════════════════════
    ss_alerts = context.get("secret_scanning", []) if context else []
    if ss_alerts and show_alert_sheets:
        ws_s = wb.create_sheet("Secret Scanning Alerts")
        ws_s.sheet_properties.tabColor = "7C3AED"
        ss_cols = ["#", "Repository", "Secret Type", "State", "Push Protection Bypassed", "Created", "URL"]
        ncols_s = len(ss_cols)
        _title_block(ws_s, "Secret Scanning Alerts", f"{len(ss_alerts)} alerts  |  {org_name}  |  {gen}", ncols_s)
        ws_s.row_dimensions[3].height = 6

        for c, h in enumerate(ss_cols, 1):
            ws_s.cell(row=4, column=c, value=h)
        _apply_header_row(ws_s, ncols_s, 4)
        ws_s.auto_filter.ref = f"A4:{get_column_letter(ncols_s)}4"
        ws_s.freeze_panes = "A5"

        for i, a in enumerate(ss_alerts):
            r = i + 5
            vals = [
                a.alert_number, a.repository.full_name,
                a.secret_type_display_name or a.secret_type or "",
                a.state.value.upper(), "YES" if a.push_protection_bypassed else "No",
                a.created_at.strftime("%Y-%m-%d") if a.created_at else "", a.html_url or "",
            ]
            for c, val in enumerate(vals, 1):
                cell = ws_s.cell(row=r, column=c, value=val)
                _apply_data_cell(cell, i)
                if c == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if c == 5 and a.push_protection_bypassed:
                    cell.font = Font(bold=True, size=10, color=_RED, name="Calibri")
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")
            ws_s.row_dimensions[r].height = 22

        _set_col_widths(ws_s, {"A": 6, "B": 28, "C": 28, "D": 10, "E": 22, "F": 12, "G": 40})

    # ══════════════════════════════════════════════════════════════════════
    #  SHEET 5 — AI Analysis
    # ══════════════════════════════════════════════════════════════════════
    ws_a = wb.create_sheet("AI Analysis")
    ws_a.sheet_properties.tabColor = _GREEN
    _title_block(ws_a, title, f"{org_name}  |  {gen}", 4)
    ws_a.row_dimensions[3].height = 6

    row = 4
    for line in content.split("\n"):
        stripped = line.strip()
        cell = ws_a.cell(row=row, column=1, value=line)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        if not stripped:
            ws_a.row_dimensions[row].height = 6
        elif stripped.startswith("#") or (stripped.isupper() and len(stripped) > 5) or re.match(r"^[=\-]{3,}$", stripped):
            cell.font = Font(bold=True, size=12, color=_NAVY, name="Calibri")
            cell.fill = PatternFill("solid", fgColor=_LIGHT_BG)
            cell.border = Border(bottom=Side(style="thin", color=_BORDER_GREY))
            ws_a.row_dimensions[row].height = 26
        elif re.match(r"^\d+[\.\)]\s", stripped):
            cell.font = Font(size=10, color=_DARK_TEXT, name="Calibri")
            cell.border = Border(left=Side(style="medium", color=_BLUE))
            ws_a.row_dimensions[row].height = 18
        elif stripped.startswith("- ") or stripped.startswith("* "):
            cell.font = Font(size=10, color=_DARK_TEXT, name="Calibri")
            ws_a.row_dimensions[row].height = 18
        else:
            cell.font = Font(size=10, color=_DARK_TEXT, name="Calibri")
            ws_a.row_dimensions[row].height = 16

        row += 1

    ws_a.column_dimensions["A"].width = 110

    # ══════════════════════════════════════════════════════════════════════
    #  Compliance template — Report Info sheet (inserted at front)
    # ══════════════════════════════════════════════════════════════════════
    if template == "compliance":
        ws_info = wb.create_sheet("Report Info", 0)
        ws_info.sheet_properties.tabColor = _RED
        _title_block(ws_info, "Report Information", "CONFIDENTIAL — FOR INTERNAL USE ONLY", 4)
        ws_info.row_dimensions[3].height = 6

        info_rows = [
            ("Report Title:", title),
            ("Organisation:", org_name),
            ("Generated:", gen),
            ("Template:", "Compliance / Audit"),
            ("Classification:", "CONFIDENTIAL — FOR INTERNAL USE ONLY"),
            ("Tool:", "AI Git Guard — GitHub Advanced Security AI Agent"),
            ("", ""),
            ("Reviewed by:", ""),
            ("Approved by:", ""),
            ("Sign-off Date:", ""),
        ]
        for i, (label, value) in enumerate(info_rows):
            r = 4 + i
            lbl_cell = ws_info.cell(row=r, column=1, value=label)
            lbl_cell.font = Font(bold=True, size=10, color=_NAVY, name="Calibri")
            lbl_cell.alignment = Alignment(vertical="center")
            val_cell = ws_info.cell(row=r, column=2, value=value)
            val_cell.font = Font(size=10, color=_DARK_TEXT, name="Calibri")
            val_cell.alignment = Alignment(vertical="center")
            if label == "Classification:":
                val_cell.font = Font(bold=True, size=10, color=_RED, name="Calibri")
            val_cell.border = Border(bottom=Side(style="thin", color=_BORDER_GREY))
            ws_info.row_dimensions[r].height = 24
        _set_col_widths(ws_info, {"A": 20, "B": 60, "C": 14, "D": 14})

    # ── Save ───────────────────────────────────────────────────────────────
    wb.active = 0
    wb.save(str(path))
    logger.info("Excel report written: %s", path)
    return path
